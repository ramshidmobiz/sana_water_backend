import random
import uuid
import json
import datetime
from customer_care.models import DiffBottlesModel
from invoice_management.models import Invoice, InvoiceDailyCollection, InvoiceItems
from van_management.models import *
from decimal import Decimal

from django.views import View
from django.db.models import Q, Sum, Count, DecimalField
from django.urls import reverse
from django.contrib import messages
from django.db import transaction, IntegrityError
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.forms import formset_factory, inlineformset_factory
from django.shortcuts import get_object_or_404, redirect, render

from .forms import  *
from .models import *
from product.models import *
from accounts.models import *
from master.functions import generate_form_errors
from master.models import RouteMaster, LocationMaster
from van_management.models import Van, Van_Routes
# from sales_management.forms import CustomerCustodyForm, ProductForm

from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Alignment
from competitor_analysis.forms import CompetitorAnalysisFilterForm
from django.db.models import Q

from van_management.models import Van, Van_Routes


def customer_custody_item(request,customer_id):
    customer_instance = Customers.objects.get(customer_id=customer_id)
    CustodyItemsFormset = formset_factory(CustodyCustomItemForm, extra=2)
    
    message = ''
    if request.method == 'POST':
        custody_custom_form = CustodyCustomForm(request.POST)
        custody_items_formset = CustodyItemsFormset(request.POST,prefix='custody_items_formset', form_kwargs={'empty_permitted': False})
        
        if custody_custom_form.is_valid() and custody_items_formset.is_valid():
            try:
                with transaction.atomic():
                    custody_custom_data = custody_custom_form.save(commit=False)
                    custody_custom_data.created_by=request.user.id
                    custody_custom_data.created_date=datetime.today()
                    custody_custom_data.modified_by=request.user.id
                    custody_custom_data.modified_date=datetime.today()
                    
                    custody_custom_data.customer=customer_instance
                    custody_custom_data.save()
                    
                    for form in custody_items_formset:
                        item = form.save(commit=False)
                        item.custody_custom = custody_custom_data
                        item.save()
                    response_data = {
                        "status": "true",
                        "title": "Successfully Created",
                        "message": "custody Item created successfully.",
                        'redirect': 'true',
                        "redirect_url": reverse('customers')
                    }
            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            message = generate_form_errors(custody_custom_form, formset=False)
            message += generate_form_errors(custody_items_formset, formset=True)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        custody_custom_form = CustodyCustomForm()
        custody_items_formset = CustodyItemsFormset(prefix='custody_items_formset')
        
        context = {
            'custody_custom_form': custody_custom_form,
            'custody_items_formset': custody_items_formset,
            'customer_instance' : customer_instance,
            
            'page_title': 'Create Custody Item',
            'page_name' : 'create Custody Item',
        }
        
        return render(request,'client_management/add_custody_items.html',context)


# def customer_custody_item(request,customer_id):
#     template_name = 'client_management/add_custody_items.html'
#     if request.method == "GET":
#         customer_exists = Customers.objects.filter(customer_id=customer_id).exists()
#         if customer_exists:
#             customer_data = Customers.objects.get(customer_id=customer_id)
#             products = ProdutItemMaster.objects.all()

#             price_list = []
#             for product in products:
#                 default_rates_exists = Product_Default_Price_Level.objects.filter(product_id=product,customer_type=customer_data.customer_type).exists()
#                 if default_rates_exists:
#                     default_rates = Product_Default_Price_Level.objects.get(product_id=product,customer_type=customer_data.customer_type)
#                     custody_items_exists = CustodyCustomItems.objects.filter(product_id=product,customer=customer_data.customer_id).exists()
#                     if custody_items_exists:
#                         custody_items = CustodyCustomItems.objects.get(product_id=product,customer=customer_data.customer_id)
#                         custody_item_id = custody_items.custody_item_id
#                         product_id = product.pk
#                         product_name = product.product_name
#                         product_rate = custody_items.rate
#                         product_count = custody_items.count
#                     else:
#                         custody_item_id = ''
#                         product_id = product.pk
#                         product_name = product.product_name
#                         product_rate = default_rates.rate
#                         product_count = 0
#                 else:
#                     custody_item_id = ''
#                     product_id = product.pk
#                     product_name = product.product_name
#                     product_rate = 0
#                     product_count = 0

#                 ite = {'custody_item_id':custody_item_id,'product_id': product_id, 'product_name': product_name,'product_rate':product_rate,'product_count':product_count}
#                 price_list.append(ite)
#         context = {'price_list': price_list,'customerid':customer_data.customer_id,'customername':customer_data.customer_name}
#         return render(request, template_name, context)
    
#     if request.method == 'POST':
#         customer_id = request.POST.get('id_customer')
#         product_ids = request.POST.getlist('price_checkbox')
#         rate = request.POST.getlist('rate')
#         count = request.POST.getlist('count')
#         id_custody_items = request.POST.getlist('id_custody_item')
#         if customer_id is not None and product_ids is not None:
#             customer_instance = Customers.objects.get(customer_id=customer_id)
#             for i, item_id in enumerate(product_ids):
#                 product_id, index = item_id.split('+')
#                 index = int(index) - 1

#                 product_instance = ProdutItemMaster.objects.get(pk=product_id)
#                 if id_custody_items[index]=='':
#                     CustodyCustomItems.objects.create(created_by=request.user,
#                                         customer=customer_instance,
#                                         rate=rate[index],
#                                         count=count[index],
#                                         product=product_instance)   
#                 else:
#                     customer_custody_instance = CustodyCustomItems.objects.get(custody_item_id=id_custody_items[index])
#                     customer_custody_instance.rate = rate[index]
#                     customer_custody_instance.count = count[index]
#                     customer_custody_instance.save()

#             messages.success(request, 'Custody Items Successfully Added.', 'alert-success')
#             return redirect('customers')
#         else:
#             messages.success(request, 'Data is not valid.', 'alert-danger')
#             context = {}
#     return render(request, template_name, context)


#ajax
def get_custody_items(request):
    if request.method == "GET":
        customer = request.GET['customer']
        if customer is not None:
            customer_exists = Customers.objects.filter(customer_id=customer).exists()
            if customer_exists:
                customer_data = Customers.objects.get(customer_id=customer)
                branch_id=request.user.branch_id.branch_id
                branch = BranchMaster.objects.get(branch_id=branch_id)
                products = Product.objects.filter(branch_id=branch)
                price_list = []
                for product in products:
                   default_rates_exists = Product_Default_Price_Level.objects.filter(product_id=product,customer_type=customer_data.customer_type).exists()
                   if default_rates_exists:
                        default_rates = Product_Default_Price_Level.objects.get(product_id=product,customer_type=customer_data.customer_type)
                        custody_items_exists = CustodyCustomItems.objects.filter(product_id=product,customer=customer_data.customer_id).exists()
                        if custody_items_exists:
                            custody_items = CustodyCustomItems.objects.get(product_id=product,customer=customer_data.customer_id)
                            custody_item_id = custody_items.custody_item_id
                            product_id = product.product_id
                            product_name = product.product_name
                            product_rate = custody_items.rate
                            product_count = custody_items.count
                        else:
                            custody_item_id = ''
                            product_id = product.product_id
                            product_name = product.product_name
                            product_rate = default_rates.rate
                            product_count = 0
                   else:
                        custody_items_exists = CustodyCustomItems.objects.filter(product_id=product,customer=customer_data.customer_id).exists()
                        if custody_items_exists:
                            custody_items = CustodyCustomItems.objects.get(product_id=product,customer=customer_data.customer_id)
                            custody_item_id = custody_items.custody_item_id
                            product_id = product.product_id
                            product_name = product.product_name
                            product_rate = custody_items.rate
                            product_count = custody_items.count
                        else:
                            custody_item_id = ''
                            product_id = product.product_id
                            product_name = product.product_name
                            product_rate = default_rates.rate
                            product_count = 0

                   ite = {'custody_item_id':custody_item_id,'product_id': product_id, 'product_name': product_name,'product_rate':product_rate,'product_count':product_count}
                   price_list.append(ite)
            dat = {'price_list': price_list}   
        return JsonResponse(dat)
    
    # --------------------------------------------------------------------------

# Vaccation
def vacation_list(request):
    template = 'client_management/vacation_list.html'
    Vacation.objects.filter(end_date__lt= date.today()).delete()
    vacation = Vacation.objects.all()
    context = {'vacation':vacation}
    return render(request, template, context)

class RouteSelection(View):
    def get(self, request):
        template = 'client_management/select_route.html'
        routes = RouteMaster.objects.all()
        return render(request, template, {'routes': routes}) 
    
class Vacation_Add(View):
    def get(self, request):
        template = 'client_management/vacation_add.html'
        form = Vacation_Add_Form
        search_form = CustomerSearchForm()
        selected_route = request.GET.get('route')
        print('root',selected_route)
        customers = Customers.objects.filter(routes = selected_route)
        search_query = request.GET.get('search_query')
        if search_query:
            customers = customers.filter(
                Q(customer_name__icontains=search_query) |
                Q(mobile_no__icontains=search_query) |
                Q(location__location_name__icontains=search_query) |
                Q(building_name__icontains=search_query) 
            )
        return render(request, template, {'form': form, 'search_form': search_form, 'customers': customers, 'selected_route':selected_route})

    def post(self, request):
        template = 'client_management/vacation_add.html'
        form = Vacation_Add_Form(request.POST)
        if form.is_valid():
            form.save()
            return redirect(vacation_list)
        return render(request, template, {'form': form})
    
class Vacation_Edit(View):
    template = 'client_management/vacation_edit.html'
    def get(self, request, vacation_id):
        vacation = Vacation.objects.get(vacation_id=vacation_id)
        form = Vacation_Edit_Form(instance=vacation)
        return render(request, self.template, {'form': form, 'vacation': vacation})

    def post(self, request, vacation_id):
        vacation = Vacation.objects.get(vacation_id=vacation_id)
        form = Vacation_Edit_Form(request.POST, instance=vacation)
        if form.is_valid():
            form.save()
            return redirect(vacation_list)
        return render(request, self.template, {'form': form, 'vacation': vacation})

class Vacation_Delete(View):
    template='client_management/vacation_delete.html'
    def get(self, request, vacation_id):
        vacation = Vacation.objects.get(vacation_id=vacation_id)
        return render(request, self.template, {'vacation':vacation})

    def post(self, request, vacation_id):
        vacation = Vacation.objects.get(vacation_id=vacation_id)
        vacation.delete()
        return redirect(vacation_list)
    


class CustomerCustodyList(View):
    template_name = 'client_management/custody_item/customer_custody_list.html'

    def get(self, request, *args, **kwargs):
        form = CompetitorAnalysisFilterForm(request.GET)
        
        user_li = CustodyCustomItems.objects.all()
        query = request.GET.get("q")
        if query:
            user_li = user_li.filter(
                Q(custody_custom__customer__customer_name__icontains=query)|
                Q(custody_custom__customer__mobile_no__icontains=query)|
                Q(custody_custom__customer__building_name__icontains=query)|
                Q(custody_custom__customer__routes__route_name__icontains=query)

            )
            

        route_filter = request.GET.get('route_name')
        if route_filter:
            user_li = user_li.filter(custody_custom__route__route_name=route_filter)

        # Fetch counts of 5 gallons, dispenser, and water cooler from Product model
        five_gallon_count = Product.objects.filter(product_name__product_name='5 Gallon').count()
        dispenser_count = Product.objects.filter(product_name__product_name='Dispenser').count()
        water_cooler_count = Product.objects.filter(product_name__product_name='Water Cooler').count()

        context = {
            'user_li': user_li,
            'form': form,
            'five_gallon_count': five_gallon_count,
            'dispenser_count': dispenser_count,
            'water_cooler_count': water_cooler_count,
        }
        return render(request, self.template_name, context)       


class AddCustodyItems(View):
    template_name = 'client_management/custody_item/add_custody_items.html'
    form_class = CustodyCustomItemForm

    def get(self, request):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = self.form_class(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            # Save the instance based on deposit_form value
            if instance.deposit_form:
                instance.amount = instance.deposit_amount
                instance.deposit_form_number = instance.deposit_number
            else:
                instance.amount = None
            instance.save()
            messages.success(request, 'Entry created successfully!')
            return redirect('add_custody_list')
        return render(request, self.template_name, {'form': form})

        
class AddCustodyList(View):
    template_name = 'client_management/custody_item/add_custody_list.html'

    def get(self, request):
        get_addedlist = CustodyCustomItems.objects.all()
        print(get_addedlist,'geddedlist')
        return render(request, self.template_name, {'get_addedlist': get_addedlist })
    
class EditCustodyItem(View):
    template_name = 'client_management/custody_item/add_custody_list.html'

    def get(self, request):
        get_addedlist = CustodyCustomItems.objects.all()
        print(get_addedlist,'geddedlist')
        return render(request, self.template_name, {'get_addedlist': get_addedlist })



class PulloutListView(View):
    template_name = 'client_management/pullout_list.html'

    # def get(self, request):
        # form = CustodyItemFilterForm(request.GET)
    def get(self, request, pk):
        customer = Customers.objects.get(customer_id=pk)
        print('customer',customer)
        custody_items = CustodyCustomItems.objects.filter(customer=customer)
        # custody_pullout_list = Customer_Custody_Items.objects.all()
        print("custody_pullout_list",list(custody_items))
        return render(request, self.template_name, {'custody_items': custody_items,'customer': customer})
    

@login_required
def customer_supply_list(request):
    """
    Customer Supply List
    :param request:
    :return: CustomerSupplys list view
    """
    filter_data = {}
    
    instances = CustomerSupply.objects.all().order_by("-created_date")
    routes = RouteMaster.objects.all()
    
    # Check if 'start_date' and 'end_date' are present in the GET request
    if request.GET.get('start_date') and request.GET.get('end_date'):
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
    else:
        start_date_str = datetime.today().date().strftime('%Y-%m-%d')
        end_date_str = datetime.today().date().strftime('%Y-%m-%d')
        
    print(start_date_str,end_date_str)
    route_name = request.GET.get('route_name')

    # Convert the string dates to date objects and filter the instances
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        instances = instances.filter(created_date__date__gte=start_date, created_date__date__lte=end_date)
        filter_data['start_date'] = start_date
        filter_data['end_date'] = end_date
    
    if route_name:
        instances = instances.filter(customer__routes__route_name=route_name)
        filter_data['route_name'] = route_name
    
    context = {
        'instances': instances,
        'page_name' : 'Customer Supply List',
        'page_title' : 'Customer Supply List',
        'routes': routes, 
        'is_customer_supply': True,
        'is_need_datetime_picker': True,
        
        'filter_data': filter_data,
    }

    return render(request, 'client_management/customer_supply/list.html', context)

@login_required
def customer_supply_info(request,pk):
    """
    Customer Supply Info
    :param request:
    :return: CustomerSupplys Info view
    """
    
    instances = CustomerSupplyItems.objects.filter(customer_supply=pk).order_by("-customer_supply__created_date")
         
    # date_range = ""
    # date_range = request.GET.get('date_range')
    # # print(date_range)

    # if date_range:
    #     start_date_str, end_date_str = date_range.split(' - ')
    #     start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
    #     end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()
    #     instances = instances.filter(date__range=[start_date, end_date])
    
    # filter_data = {}
    # query = request.GET.get("q")
    
    # if query:

    #     instances = instances.filter(
    #         Q(customer_supply_no__icontains=query) |
    #         Q(product__customer_supply_id__icontains=query) 
    #     )
    #     title = "Customer Supply List - %s" % query
    #     filter_data['q'] = query
    
    context = {
        'instances': instances,
        'page_name' : 'Customer Supply List',
        'page_title' : 'Customer Supply List',
        # 'filter_data' :filter_data,
        # 'date_range': date_range,
        
        'is_customer_supply': True,
        'is_need_datetime_picker': True,
    }

    return render(request, 'client_management/customer_supply/info.html', context)

@login_required
def customer_supply_customers(request,pk):
    filter_data = {}
    
    instances = Customers.objects.all()
    
    if request.GET.get('route'):
        instances = instances.filter(routes__pk=request.GET.get('route'))
        
    if request.GET.get('building_no'):
        instances = instances.filter(door_house_no=request.GET.get('building_no'))
    
    query = request.GET.get("q")
    
    if query:

        instances = instances.filter(
            Q(customer_id__icontains=query) |
            Q(mobile_no__icontains=query) |
            Q(whats_app__icontains=query) |
            Q(customer_name__icontains=query) 
        )
        title = "Customer Supply Customers - %s" % query
        filter_data['q'] = query
        
    route_instances = RouteMaster.objects.all()

    context = {
        'instances': instances,
        'route_instances' : route_instances,
        
        'page_title': 'Create customer supply',
        'customer_supply_page': True,
        'is_need_datetime_picker': True
    }
    
    return render(request,'client_management/customer_supply/customer_list.html',context)

def create_customer_supply(request):
    
    message = ''
    if request.method == 'POST':
        customer_supply_form = CustomerSupplyForm(request.POST)
        customer_supply_items_form = CustomerSupplyItemsForm(request.POST)
        
        if customer_supply_form.is_valid() and customer_supply_items_form.is_valid():
            try:
                with transaction.atomic():
                    customer_supply = customer_supply_form.save(commit=False)
                    customer_supply.created_by = str(request.user.id)
                    customer_supply.save()
                    
                    data = customer_supply_items_form.save(commit=False)
                    data.customer_supply = customer_supply
                    data.save()
                    
                    if (update_customer_stock:=CustomerSupplyStock.objects.filter(customer=customer_supply.customer,product=data.product)).exists():
                            stock = update_customer_stock.first()
                            stock.stock_quantity += data.quantity
                            stock.save()
                    else:
                        CustomerSupplyStock.objects.create(
                            product=data.product,
                            customer=customer_supply.customer,
                            stock_quantity = data.quantity,
                        )
                    
                    response_data = {
                        "status": "true",
                        "title": "Successfully Created",
                        "message": "Customer Supply created successfully.",
                        'redirect': 'true',
                        "redirect_url": reverse('customer_supply:customer_supply_list')
                    }
                    
            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
            message = generate_form_errors(customer_supply_form,formset=False)
            message += generate_form_errors(customer_supply_items_form,formset=False)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        customer_supply_form = CustomerSupplyForm()
        customer_supply_items_form = CustomerSupplyItemsForm()
        
        context = {
            'customer_supply_form': customer_supply_form,
            'customer_supply_items_form': customer_supply_items_form,
            
            'page_title': 'Create customer supply',
            'customer_supply_page': True,
            'is_need_datetime_picker': True
        }
        
        return render(request,'client_management/customer_supply/create.html',context)


@login_required
def edit_customer_supply(request,pk):
    """
    edit operation of customer_supply
    :param request:
    :param pk:
    :return:
    """
    message = ''
    customer_supply_instance = CustomerSupply.objects.get(pk=pk)
    supply_items_instances = CustomerSupplyItems.objects.filter(customer_supply=customer_supply_instance)
    
    if supply_items_instances.exists():
        extra = 0
    else:
        extra = 1 

    SupplyItemsFormset = inlineformset_factory(
        CustomerSupply,
        CustomerSupplyItems,
        extra=extra,
        form=CustomerSupplyItemsForm,
    )
    
    if request.method == 'POST':
        customer_supply_form = CustomerSupplyForm(request.POST,instance=customer_supply_instance)
        customer_supply_items_formset = SupplyItemsFormset(request.POST,request.FILES,
                                            instance=customer_supply_instance,
                                            prefix='customer_supply_items_formset',
                                            form_kwargs={'empty_permitted': False}) 
        
        if customer_supply_form.is_valid() and  customer_supply_items_formset.is_valid() :
            
            # #create
            five_gallon_qty = supply_items_instances.filter(product__product_name="5 Gallon").aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
            DiffBottlesModel.objects.filter(
                delivery_date__date=customer_supply_instance.created_date.date(),
                assign_this_to=customer_supply_instance.salesman,
                customer=customer_supply_instance.customer_id
                ).update(status='pending')
            
            invoice_instance = Invoice.objects.get(created_date__date=customer_supply_instance.created_date.date(),customer=customer_supply_instance.customer,reference_no=customer_supply_instance.reference_number)
            invoice_items_instances = InvoiceItems.objects.filter(invoice=invoice_instance)
            InvoiceDailyCollection.objects.filter(
                invoice=invoice_instance,
                created_date__date=customer_supply_instance.created_date.date(),
                customer=customer_supply_instance.customer,
                salesman=customer_supply_instance.salesman
                ).delete()
            invoice_items_instances.delete()
            invoice_instance.delete()
            
            balance_amount = customer_supply_instance.subtotal - customer_supply_instance.amount_recieved
            if customer_supply_instance.amount_recieved < customer_supply_instance.subtotal:
                OutstandingAmount.objects.filter(
                    customer_outstanding__product_type="amount",
                    customer_outstanding__customer=customer_supply_instance.customer,
                    customer_outstanding__created_by=customer_supply_instance.salesman.pk,
                    customer_outstanding__created_date=customer_supply_instance.created_date,
                    amount=balance_amount
                ).delete()
                
                customer_outstanding_report_instance=CustomerOutstandingReport.objects.get(customer=customer_supply_instance.customer,product_type="amount")
                customer_outstanding_report_instance.value -= Decimal(balance_amount)
                customer_outstanding_report_instance.save()
                
            elif customer_supply_instance.amount_recieved > customer_supply_instance.subtotal:
                OutstandingAmount.objects.filter(
                    customer_outstanding__product_type="amount",
                    customer_outstanding__customer=customer_supply_instance.customer,
                    customer_outstanding__created_by=customer_supply_instance.salesman.pk,
                    customer_outstanding__created_date=customer_supply_instance.created_date,
                    amount=balance_amount
                ).delete()
                
                customer_outstanding_report_instance=CustomerOutstandingReport.objects.get(customer=customer_supply_instance.customer,product_type="amount")
                customer_outstanding_report_instance.value += Decimal(balance_amount)
                customer_outstanding_report_instance.save()
                
            if (digital_coupons_instances:=CustomerSupplyDigitalCoupon.objects.filter(customer_supply=customer_supply_instance)).exists():
                digital_coupons_instance = digital_coupons_instances.first()
                CustomerCouponStock.objects.get(
                    coupon_method="digital",
                    customer=customer_supply_instance.customer,
                    coupon_type_id__coupon_type_name="Other"
                    ).count += digital_coupons_instance.count
            
            elif (manual_coupon_instances := CustomerSupplyCoupon.objects.filter(customer_supply=customer_supply_instance)).exists():
                manual_coupon_instance = manual_coupon_instances.first()
                leaflets_to_update = manual_coupon_instance.leaf.filter(used=True)
                updated_count = leaflets_to_update.count()

                if updated_count > 0:
                    first_leaflet = leaflets_to_update.first()

                    if first_leaflet and CustomerCouponStock.objects.filter(
                            customer=customer_supply_instance.customer,
                            coupon_method="manual",
                            coupon_type_id=first_leaflet.coupon.coupon_type
                        ).exists():
                        # Update the CustomerCouponStock
                        customer_stock_instance = CustomerCouponStock.objects.get(
                            customer=customer_supply_instance.customer,
                            coupon_method="manual",
                            coupon_type_id=first_leaflet.coupon.coupon_type
                        )
                        customer_stock_instance.count += Decimal(updated_count)
                        customer_stock_instance.save()
                        
                        if five_gallon_qty < Decimal(customer_supply_instance.collected_empty_bottle) :
                            balance_empty_bottle = Decimal(customer_supply_instance.collected_empty_bottle) - five_gallon_qty
                            if CustomerOutstandingReport.objects.filter(customer=customer_supply_instance.customer,product_type="emptycan").exists():
                                outstanding_instance = CustomerOutstandingReport.objects.get(customer=customer_supply_instance.customer,product_type="emptycan")
                                outstanding_instance.value += Decimal(balance_empty_bottle)
                                outstanding_instance.save()
                                
                        elif five_gallon_qty > Decimal(customer_supply_instance.collected_empty_bottle) :
                            balance_empty_bottle = five_gallon_qty - Decimal(customer_supply_instance.collected_empty_bottle)
                            
                            outstanding_instance = CustomerOutstanding.objects.filter(
                                product_type="emptycan",
                                created_by=customer_supply_instance.salesman.pk,
                                customer=customer_supply_instance.customer,
                                created_date=customer_supply_instance.created_date,
                            ).first()

                            outstanding_product = OutstandingProduct.objects.filter(
                                empty_bottle=balance_empty_bottle,
                                customer_outstanding=outstanding_instance,
                            )
                            outstanding_instance = {}

                            try:
                                outstanding_instance=CustomerOutstandingReport.objects.get(customer=customer_supply_instance.customer,product_type="emptycan")
                                outstanding_instance.value -= Decimal(outstanding_product.aggregate(total_empty_bottle=Sum('empty_bottle'))['total_empty_bottle'])
                                outstanding_instance.save()
                            except:
                                pass
                        leaflets_to_update.update(used=False)
                        outstanding_product.delete()
                        
            for item_data in supply_items_instances:
                if VanProductStock.objects.filter(created_date=customer_supply_instance.created_date.date(),product=item_data.product,van__salesman=customer_supply_instance.salesman).exists():
                    if item_data.product.product_name == "5 Gallon" :
                        # total_fivegallon_qty -= Decimal(five_gallon_qty)
                        if VanProductStock.objects.filter(created_date=customer_supply_instance.created_date.date(),product=item_data.product,van__salesman=customer_supply_instance.salesman).exists():
                            empty_bottle = VanProductStock.objects.get(
                                product=item_data.product,
                                created_date=customer_supply_instance.created_date.date(),
                                van__salesman=customer_supply_instance.salesman,
                            )
                            empty_bottle.empty_can_count -= customer_supply_instance.collected_empty_bottle
                            empty_bottle.save()
                        
                    vanstock = VanProductStock.objects.get(created_date=customer_supply_instance.created_date.date(),product=item_data.product,van__salesman=customer_supply_instance.salesman)
                    vanstock.stock += item_data.quantity
                    vanstock.save()
            
            customer_supply_instance.delete()
            supply_items_instances.delete()
                
            response_data = {
                "status": "true",
                "title": "Successfully Updated",
                "message": "customer supply Updated Successfully.",
                'redirect': 'true',
                "redirect_url": reverse('customer_supply:customer_supply_list'),
                "return" : True,
            }
    
        else:
            message = generate_form_errors(customer_supply_form,formset=False)
            message += generate_form_errors(customer_supply_items_formset,formset=True)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
                        
    else:
        customer_supply_form = CustomerSupplyForm(instance=customer_supply_instance)
        customer_supply_items_formset = SupplyItemsFormset(queryset=supply_items_instances,
                                                    prefix='customer_supply_items_formset',
                                                    instance=customer_supply_instance)

        context = {
            'customer_supply_form': customer_supply_form,
            'customer_supply_items_formset': customer_supply_items_formset,
            'customer_instance': customer_supply_instance.customer,
            
            'message': message,
            'page_name' : 'edit customer supply',
            'customer_supply_page': True,   
            'is_edit' : True,        
        }

        return render(request, 'client_management/customer_supply/create.html', context)
    
@login_required
def delete_customer_supply(request, pk):
    """
    customer_supply deletion, it only marks as is_deleted field to true
    :param request:
    :param pk:
    :return:
    """
    try:
        with transaction.atomic():
            customer_supply_instance = get_object_or_404(CustomerSupply, pk=pk)
            supply_items_instances = CustomerSupplyItems.objects.filter(customer_supply=customer_supply_instance)
            five_gallon_qty = supply_items_instances.filter(product__product_name="5 Gallon").aggregate(total_quantity=Sum('quantity', output_field=DecimalField()))['total_quantity'] or 0
            
            DiffBottlesModel.objects.filter(
                delivery_date__date=customer_supply_instance.created_date.date(),
                assign_this_to=customer_supply_instance.salesman,
                customer=customer_supply_instance.customer_id
                ).update(status='pending')
            
            # Handle invoice related deletions
            handle_invoice_deletion(customer_supply_instance)
            
            # Handle outstanding amount adjustments
            handle_outstanding_amounts(customer_supply_instance, five_gallon_qty)
            
            # Handle coupon deletions and adjustments
            handle_coupons(customer_supply_instance, five_gallon_qty)
            
            # Update van product stock and empty bottle counts
            update_van_product_stock(customer_supply_instance, supply_items_instances, five_gallon_qty)
            
            # Mark customer supply and items as deleted
            customer_supply_instance.delete()
            supply_items_instances.delete()
            
            response_data = {
                "status": "true",
                "title": "Successfully Deleted",
                "message": "Customer supply successfully deleted.",
                "reload": "true",
            }
            
            return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    except Exception as e:
        response_data = {
            "status": "false",
            "title": "Deletion Failed",
            "message": str(e),
        }
        return HttpResponse(json.dumps(response_data), content_type='application/javascript')


def handle_invoice_deletion(customer_supply_instance):
    if Invoice.objects.filter(created_date__date=customer_supply_instance.created_date.date(), customer=customer_supply_instance.customer, reference_no=customer_supply_instance.reference_number).exists():
        invoice_instance = Invoice.objects.get(created_date__date=customer_supply_instance.created_date.date(), customer=customer_supply_instance.customer, reference_no=customer_supply_instance.reference_number)
        invoice_items_instances = InvoiceItems.objects.filter(invoice=invoice_instance)
        
        InvoiceDailyCollection.objects.filter(
            invoice=invoice_instance,
            created_date__date=customer_supply_instance.created_date.date(),
            customer=customer_supply_instance.customer,
            salesman=customer_supply_instance.salesman
            ).delete()
        
        invoice_items_instances.delete()
        invoice_instance.delete()


def handle_outstanding_amounts(customer_supply_instance, five_gallon_qty):
    balance_amount = customer_supply_instance.subtotal - customer_supply_instance.amount_recieved
    
    if customer_supply_instance.amount_recieved < customer_supply_instance.subtotal:
        OutstandingAmount.objects.filter(
            customer_outstanding__product_type="amount",
            customer_outstanding__customer=customer_supply_instance.customer,
            customer_outstanding__created_by=customer_supply_instance.salesman.pk,
            customer_outstanding__created_date=customer_supply_instance.created_date,
            amount=balance_amount
        ).delete()
        
        if CustomerOutstandingReport.objects.filter(customer=customer_supply_instance.customer, product_type="amount").exists():
            customer_outstanding_report_instance = CustomerOutstandingReport.objects.get(customer=customer_supply_instance.customer, product_type="amount")
            customer_outstanding_report_instance.value -= Decimal(balance_amount)
            customer_outstanding_report_instance.save()
        
    elif customer_supply_instance.amount_recieved > customer_supply_instance.subtotal:
        OutstandingAmount.objects.filter(
            customer_outstanding__product_type="amount",
            customer_outstanding__customer=customer_supply_instance.customer,
            customer_outstanding__created_by=customer_supply_instance.salesman.pk,
            customer_outstanding__created_date=customer_supply_instance.created_date,
            amount=balance_amount
        ).delete()
        
        customer_outstanding_report_instance = CustomerOutstandingReport.objects.get(customer=customer_supply_instance.customer, product_type="amount")
        customer_outstanding_report_instance.value += Decimal(balance_amount)
        customer_outstanding_report_instance.save()


def handle_coupons(customer_supply_instance, five_gallon_qty):
    if (digital_coupons_instances := CustomerSupplyDigitalCoupon.objects.filter(customer_supply=customer_supply_instance)).exists():
        digital_coupons_instance = digital_coupons_instances.first()
        CustomerCouponStock.objects.filter(
            coupon_method="digital",
            customer=customer_supply_instance.customer,
            coupon_type_id__coupon_type_name="Other"
        ).update(count=F('count') + digital_coupons_instance.count)
    
    elif (manual_coupon_instances := CustomerSupplyCoupon.objects.filter(customer_supply=customer_supply_instance)).exists():
        manual_coupon_instance = manual_coupon_instances.first()
        leaflets_to_update = manual_coupon_instance.leaf.filter(used=True)
        updated_count = leaflets_to_update.count()

        if updated_count > 0:
            first_leaflet = leaflets_to_update.first()

            if first_leaflet and CustomerCouponStock.objects.filter(
                    customer=customer_supply_instance.customer,
                    coupon_method="manual",
                    coupon_type_id=first_leaflet.coupon.coupon_type
                ).exists():
                customer_stock_instance = CustomerCouponStock.objects.get(
                    customer=customer_supply_instance.customer,
                    coupon_method="manual",
                    coupon_type_id=first_leaflet.coupon.coupon_type
                )
                customer_stock_instance.count += Decimal(updated_count)
                customer_stock_instance.save()
                
                handle_empty_bottle_outstanding(customer_supply_instance, five_gallon_qty)
                
                leaflets_to_update.update(used=False)


def handle_empty_bottle_outstanding(customer_supply_instance, five_gallon_qty):
    if five_gallon_qty < Decimal(customer_supply_instance.collected_empty_bottle):
        balance_empty_bottle = Decimal(customer_supply_instance.collected_empty_bottle) - five_gallon_qty
        if CustomerOutstandingReport.objects.filter(customer=customer_supply_instance.customer, product_type="emptycan").exists():
            outstanding_instance = CustomerOutstandingReport.objects.get(customer=customer_supply_instance.customer, product_type="emptycan")
            outstanding_instance.value += Decimal(balance_empty_bottle)
            outstanding_instance.save()
            
    elif five_gallon_qty > Decimal(customer_supply_instance.collected_empty_bottle):
        balance_empty_bottle = five_gallon_qty - Decimal(customer_supply_instance.collected_empty_bottle)
        
        outstanding_instance = CustomerOutstanding.objects.filter(
            product_type="emptycan",
            created_by=customer_supply_instance.salesman.pk,
            customer=customer_supply_instance.customer,
            created_date=customer_supply_instance.created_date,
        ).first()

        outstanding_product = OutstandingProduct.objects.filter(
            empty_bottle=balance_empty_bottle,
            customer_outstanding=outstanding_instance,
        )
        outstanding_instance = {}

        try:
            outstanding_instance = CustomerOutstandingReport.objects.get(customer=customer_supply_instance.customer, product_type="emptycan")
            outstanding_instance.value -= Decimal(outstanding_product.aggregate(total_empty_bottle=Sum('empty_bottle'))['total_empty_bottle'])
            outstanding_instance.save()
        except:
            pass
        outstanding_product.delete()


def update_van_product_stock(customer_supply_instance, supply_items_instances, five_gallon_qty):
    for item_data in supply_items_instances:
        if VanProductStock.objects.filter(created_date=customer_supply_instance.created_date.date(), product=item_data.product, van__salesman=customer_supply_instance.salesman).exists():
            if item_data.product.product_name == "5 Gallon":
                empty_bottle = VanProductStock.objects.get(
                    product=item_data.product,
                    created_date=customer_supply_instance.created_date.date(),
                    van__salesman=customer_supply_instance.salesman,
                )
                empty_bottle.empty_can_count -= customer_supply_instance.collected_empty_bottle
                empty_bottle.save()
            
            vanstock = VanProductStock.objects.get(created_date=customer_supply_instance.created_date.date(), product=item_data.product, van__salesman=customer_supply_instance.salesman)
            vanstock.stock += item_data.quantity
            vanstock.save()


#------------------------------REPORT----------------------------------------

def client_report(request):
    instances = Customers.objects.order_by('-created_date')  # Order by latest created date
    return render(request, 'client_management/client_report.html', {'instances': instances})


def clientdownload_pdf(request, customer_id):
    customer = get_object_or_404(Customers, pk=customer_id)
    template_path = 'client_management/client_report_pdf.html'
    context = {'customer': customer}

    # Logic to generate PDF for the specific customer
    pdf_content = f"PDF content for {customer.customer_name}"
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{customer.customer_name}_report.pdf"'
    template = get_template(template_path)
    html = template.render(context)

    # Create PDF
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')

    return response


def clientexport_to_csv(request, customer_id):
    customer = get_object_or_404(Customers, pk=customer_id)

    # Create an Excel workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # Set the title
    title_cell = ws.cell(row=1, column=1, value="Client Details")
    title_cell.alignment = Alignment(horizontal='center')  # Align the title to the center
    ws.merge_cells('A1:F1')  # Merge cells for the title

    # Define data to be written
    data = [
        ['Customer Name:', customer.customer_name],
        ['Address:', f"{customer.building_name} {customer.door_house_no}"],
        ['Contact:', customer.mobile_no],
        ['Customer Type:', customer.customer_type],
        ['Sales Type:', customer.sales_type],
    ]

    # Write data to the worksheet
    for row in data:
        ws.append(row)

    # Set response headers for Excel file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{customer.customer_name}_report.xlsx"'

    # Save the workbook to the HttpResponse
    wb.save(response)

    return response


def custody_items_list_report(request):
    if request.method == 'GET':
        
        instances = CustodyCustom.objects.all()
        # if start_date_str and end_date_str:
        #     start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        #     end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        #     instances = CustodyCustomItems.objects.filter(custody_custom__created_date__range=[start_date, end_date])
        # else:
        
        return render(request, 'client_management/custody_items_list_report.html', {'instances': instances})


def custody_issue(request):
    instances = CustodyCustom.objects.all().order_by("-created_date")

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        instances = instances.filter(created_date__range=[start_date, end_date])
    
        
    customer_product_counts = {}

    for instance in instances:
        customer = instance.customer
        print(customer,"customer")
        if customer not in customer_product_counts:
            customer_product_counts[customer] = {
                '5_gallon_deposit': 0,
                '5_gallon_non_deposit': 0,
                'dispenser_deposit': 0,
                'dispenser_non_deposit': 0,
                'cooler_deposit': 0,
                'cooler_non_deposit': 0,
            }

        custody_items = instance.custodycustomitems_set.all()

        for custody_item in custody_items:
            product = custody_item.product
            if product:
                if product.product_name == '5 Gallon':
                    if custody_item.custody_custom.deposit_type == 'deposit':
                        customer_product_counts[customer]['5_gallon_deposit'] += 1
                    else:
                        customer_product_counts[customer]['5_gallon_non_deposit'] += 1
                elif product.product_name == 'Dispenser':
                    if custody_item.custody_custom.deposit_type == 'deposit':
                        customer_product_counts[customer]['dispenser_deposit'] += 1
                    else:
                        customer_product_counts[customer]['dispenser_non_deposit'] += 1
                elif product.product_name == 'Cooler':
                    if custody_item.custody_custom.deposit_type == 'deposit':
                        customer_product_counts[customer]['cooler_deposit'] += 1
                    else:
                        customer_product_counts[customer]['cooler_non_deposit'] += 1

    for customer, counts in customer_product_counts.items():
        for key, value in counts.items():
            if value == 0:
                customer_product_counts[customer][key] = '--'

    return render(request, 'client_management/custody_issue.html', {'customer_product_counts': customer_product_counts})



def get_customercustody(request, customer_id):
    customer = Customers.objects.get(customer_id=customer_id)  # Use get() if customer_id is unique
    print(customer, "customer")
    
    custody_items = CustodyCustom.objects.filter(customer=customer)
    print("custody_items", custody_items)
    
    custody_items_with_products = []
    
    for custody_item in custody_items:
        custody_custom_items = CustodyCustomItems.objects.filter(custody_custom=custody_item)
        custody_item_data = {
            'custody_custom': custody_item,
            'custody_custom_items': custody_custom_items,
            'products': [item.product for item in custody_custom_items]  
        }
        custody_items_with_products.append(custody_item_data)

    context = {'custody_items_with_products': custody_items_with_products,'customer':customer}
    return render(request, 'client_management/customer_custody_items.html', context)

def custody_report(request):
    instances = CustodyCustom.objects.all().order_by("-created_date")

    start_date_str = request.GET.get('start_date')
    print("start_date_str",    start_date_str)

    end_date_str = request.GET.get('end_date')
    print("end_date_str",end_date_str)
    
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        instances = instances.filter(created_date__range=[start_date, end_date])

    custody_items_with_products = []
    
    for custody_item in instances:
        custody_custom_items = CustodyCustomItems.objects.filter(custody_custom=custody_item)
        print("custody_custom_items",custody_custom_items)
        custody_item_data = {
            'custody_custom': custody_item,
            'custody_custom_items': custody_custom_items,
            'products': [item.product for item in custody_custom_items]  
        }
        custody_items_with_products.append(custody_item_data)

    context = {'custody_items_with_products': custody_items_with_products,'instances':instances}
    print("context",context)
    return render(request, 'client_management/custody_report.html', context)


class CouponCountList(View):
    template_name = 'client_management/coupon_count_list.html'

    def get(self, request, pk, *args, **kwargs):
        customer = Customers.objects.get(customer_id=pk)
        customers = CustomerCouponStock.objects.filter(customer=customer)

        # Calculate total count
        total_count = customers.aggregate(total_count=Sum('count'))['total_count'] or 0

        context = {
            'customers': customers,
            'pk': pk,  # Pass pk to the template context
            'total_count': total_count,  # Pass total count to the template context
        }

        return render(request, self.template_name, context)







#
# class CouponCountList(View):
#     template_name = 'client_management/coupon_count_list.html'
#
#     def get(self, request, pk, *args, **kwargs):
#         customer = Customers.objects.get(customer_id=pk)
#         customers = CustomerCouponStock.objects.filter(customer=customer)
#
#         context = {
#             'customers': customers,
#             'pk': pk,  # Pass pk to the template context
#         }
#
#         return render(request, self.template_name, context)
#
#     def post(self, request, pk, *args, **kwargs):
#         customer = Customers.objects.get(customer_id=pk)
#         coupon_code = request.POST.get('coupon_code')
#         CustomerCouponStock.objects.create(customer=customer, coupon_code=coupon_code)
#         return redirect('coupon_count_list', pk=pk)



#
# def edit_coupon_count(request, pk):
#     customer_coupon_stock = get_object_or_404(CustomerCouponStock, customer_id=pk)
#     form = CoupenEditForm(instance=customer_coupon_stock)
#
#     if request.method == 'POST':
#         form = CoupenEditForm(request.POST, instance=customer_coupon_stock)
#         if form.is_valid():
#             form.save()
#             messages.success(request, 'Customer Coupon Stock Updated successfully!')
#             return redirect('customers')
#         else:
#             messages.error(request, 'Invalid form data. Please check the input.')
#
#     return render(request, 'client_management/edit_coupon_count.html', {'form': form})

def new_coupon_count(request,pk):
    if request.method == 'POST':
        form = CoupenEditForm(request.POST)
        if form.is_valid():
            # data = form.save(commit=False)
            # data.customer = Customers.objects.get(pk=pk)
            # data.save()
            coupon_type_id = CouponType.objects.get(pk=form.cleaned_data['coupon_type_id'].pk)
            
            coupon_method = "manual"
            if coupon_type_id.coupon_type_name is "Other":
                coupon_method = "digital"
            try:
                data = CustomerCouponStock.objects.get(
                    customer_id=Customers.objects.get(pk=pk),
                    coupon_type_id=coupon_type_id,
                    coupon_method=coupon_method
                )
            except CustomerCouponStock.DoesNotExist:
                data = CustomerCouponStock.objects.create(
                    customer_id=Customers.objects.get(pk=pk),
                    coupon_type_id=coupon_type_id,
                    coupon_method=coupon_method,
                    count=0
                )

            data.count += Decimal(form.cleaned_data['count'])
            data.save()

            messages.success(request, 'New coupon count added successfully!')
            return redirect('coupon_count_list' ,data.customer_id)
        else:
            messages.error(request, 'Invalid form data. Please check the input.')
    else:
        form = CoupenEditForm()

    return render(request, 'client_management/edit_coupon_count.html', {'form': form})

def delete_count(request, pk):
    customer_coupon_stock = get_object_or_404(CustomerCouponStock, pk=pk)

    if request.method == 'POST':
        customer_pk = customer_coupon_stock.customer.pk
        customer_coupon_stock.delete()
        messages.success(request, 'Coupon count deleted successfully!')
        return redirect('coupon_count_list', pk=customer_pk)

    return redirect('coupon_count_list')

@login_required
def customer_outstanding_list(request):
    """
    Customer Outstanding List
    :param request:
    :return: Customer Outstanding list view
    """
    filter_data = {}
    reports = CustomerOutstandingReport.objects.all()
    
    query = request.GET.get("q")
    
    if query:

        reports = reports.filter(
            Q(customer__customer_name__icontains=query) |
            Q(customer__customer_id__icontains=query) |
            Q(customer__mobile_no__icontains=query) |
            Q(customer__whats_app__icontains=query) |
            Q(customer__email_id__icontains=query) |
            Q(building_name__invoice_id__icontains=query) 
        )
        title = "Outstanding List - %s" % query
        filter_data['q'] = query
    
    route_filter = request.GET.get('route_name')
    if route_filter:
            reports = reports.filter(customer__routes__route_name=route_filter)
    route_li = RouteMaster.objects.all()
    if request.GET.get("customer_pk"):
        reports = reports.filter(customer__pk=request.GET.get("customer_pk"))

    # Organize data for rendering in the template
    customer_data = {}
    for report in reports:
        customer_id = report.customer.pk
        if customer_id not in customer_data:
            customer_data[customer_id] = {
                'customer': report.customer,
                'amount': 0,
                'empty_can': 0,
                'coupons': 0
            }
        
        # Add product values based on product type
        if report.product_type == 'amount':
            customer_data[customer_id]['amount'] += report.value
        elif report.product_type == 'emptycan':
            customer_data[customer_id]['empty_can'] += report.value
        elif report.product_type == 'coupons':
            customer_data[customer_id]['coupons'] += report.value
            
    context = {
        'instances': customer_data.values(),
        'page_name' : 'Customer Outstanding List',
        'page_title' : 'Customer Outstanding List',
        'customer_pk': request.GET.get("customer_pk"),
        
        'is_customer_outstanding': True,
        'is_need_datetime_picker': True,
        'route_li': route_li,
        'filter_data': filter_data,
    }

    return render(request, 'client_management/customer_outstanding/list.html', context)

@login_required
def create_customer_outstanding(request):
    customer_pk = request.GET.get("customer_pk")
    
    message = ''
    if request.method == 'POST':
        if not customer_pk :
            customer_outstanding_form = CustomerOutstandingForm(request.POST)
        else:
            customer_outstanding_form = CustomerOutstandingSingleForm(request.POST)
            
        customer_outstanding_amount_form = CustomerOutstandingAmountForm(request.POST)
        customer_outstanding_bottles_form = CustomerOutstandingBottleForm(request.POST)
        customer_outstanding_coupon_form = CustomerOutstandingCouponsForm(request.POST)
        
        is_form_valid = False
        if request.POST.get('product_type') == "amount":
            if customer_outstanding_form.is_valid() and customer_outstanding_amount_form.is_valid():
                is_form_valid = True
            else:
                message = generate_form_errors(customer_outstanding_form,formset=False)
                message += generate_form_errors(customer_outstanding_amount_form,formset=False)
                
        if request.POST.get('product_type') == "emptycan":
            if customer_outstanding_form.is_valid() and customer_outstanding_bottles_form.is_valid():
                is_form_valid = True
            else:
                message = generate_form_errors(customer_outstanding_form,formset=False)
                message += generate_form_errors(customer_outstanding_bottles_form,formset=False)
                
        if request.POST.get('product_type') == "coupons":
            if customer_outstanding_form.is_valid() and customer_outstanding_coupon_form.is_valid():
                is_form_valid = True
            else:
                message = generate_form_errors(customer_outstanding_form,formset=False)
                message += generate_form_errors(customer_outstanding_coupon_form,formset=False)
            
        if is_form_valid :
            try:
                with transaction.atomic():
                    # Save customer_outstanding_form data
                    outstanding_data = customer_outstanding_form.save(commit=False)
                    outstanding_data.created_by = request.user.id
                    outstanding_data.created_date = datetime.today()
                    if customer_pk :
                        print("custo_pk")
                        outstanding_data.customer = Customers.objects.get(pk=customer_pk)
                    outstanding_data.save()
                    
                    # Save data based on product type
                    if outstanding_data.product_type == "amount":
                        outstanding_amount = customer_outstanding_amount_form.save(commit=False)
                        outstanding_amount.customer_outstanding = outstanding_data
                        outstanding_amount.save()
                        
                        # Check if there is an existing report entry
                        existing_report = CustomerOutstandingReport.objects.filter(
                            customer=outstanding_data.customer,
                            product_type='amount'
                        ).first()
                        
                        if existing_report:
                            existing_report.value += outstanding_amount.amount
                            existing_report.save()
                        else:
                            CustomerOutstandingReport.objects.create(
                                product_type='amount',
                                value=outstanding_amount.amount,
                                customer=outstanding_data.customer
                            )
                            
                        random_part = str(random.randint(1000, 9999))
                        invoice_number = f'WTR-{random_part}'
                        
                        # Create the invoice
                        invoice = Invoice.objects.create(
                            invoice_no=invoice_number,
                            created_date=datetime.today(),
                            net_taxable=outstanding_amount.amount,
                            vat=0,
                            discount=0,
                            amout_total=outstanding_amount.amount,
                            amout_recieved=0,
                            customer=outstanding_amount.customer_outstanding.customer,
                            reference_no="oustading added for customer"
                        )
                        
                        if outstanding_amount.customer_outstanding.customer.sales_type == "CREDIT":
                            invoice.invoice_type = "credit_invoive"
                            invoice.save()

                        # Create invoice items
                        item = ProdutItemMaster.objects.get(product_name="5 Gallon")
                        InvoiceItems.objects.create(
                            category=item.category,
                            product_items=item,
                            qty=0,
                            rate=outstanding_amount.customer_outstanding.customer.rate,
                            invoice=invoice,
                            remarks='invoice genereted from backend reference no : ' + invoice.reference_no
                        )

                    
                    elif outstanding_data.product_type == "emptycan":
                        outstanding_bottle = customer_outstanding_bottles_form.save(commit=False)
                        outstanding_bottle.customer_outstanding = outstanding_data
                        outstanding_bottle.save()
                        
                        # Similar logic for empty can
                        # Check if there is an existing report entry
                        existing_report = CustomerOutstandingReport.objects.filter(
                            customer=outstanding_data.customer,
                            product_type='emptycan'
                        ).first()
                        
                        if existing_report:
                            existing_report.value += outstanding_bottle.empty_bottle
                            existing_report.save()
                        else:
                            CustomerOutstandingReport.objects.create(
                                product_type='emptycan',
                                value=outstanding_bottle.empty_bottle,
                                customer=outstanding_data.customer
                            )
                        
                    elif outstanding_data.product_type == "coupons":
                        outstanding_coupon = customer_outstanding_coupon_form.save(commit=False)
                        outstanding_coupon.customer_outstanding = outstanding_data
                        outstanding_coupon.save()
                        
                        # Similar logic for coupons
                        # Check if there is an existing report entry
                        existing_report = CustomerOutstandingReport.objects.filter(
                            customer=outstanding_data.customer,
                            product_type='coupons'
                        ).first()
                        
                        if existing_report:
                            existing_report.value += outstanding_coupon.count
                            existing_report.save()
                        else:
                            CustomerOutstandingReport.objects.create(
                                product_type='coupons',
                                value=outstanding_coupon.count,
                                customer=outstanding_data.customer
                            ) 
                                        
                    if not customer_pk:
                        redirect_url = reverse('customer_outstanding_list')
                    else:
                        redirect_url = reverse('customer_outstanding_list') + f'?customer_pk={customer_pk}'
                        
                    response_data = {
                        "status": "true",
                        "title": "Successfully Created",
                        "message": "Customer Supply created successfully.",
                        'redirect': 'true',
                        'redirect_url': redirect_url,
                    }
                    
            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        if not customer_pk :
            customer_outstanding_form = CustomerOutstandingForm()
        else:
            customer_outstanding_form = CustomerOutstandingSingleForm()
        
        customer_outstanding_amount_form = CustomerOutstandingAmountForm()
        customer_outstanding_bottles_form = CustomerOutstandingBottleForm()
        customer_outstanding_coupon_form = CustomerOutstandingCouponsForm()
        
        context = {
            'customer_outstanding_form': customer_outstanding_form,
            'customer_outstanding_amount_form': customer_outstanding_amount_form,
            'customer_outstanding_bottles_form': customer_outstanding_bottles_form,
            'customer_outstanding_coupon_form': customer_outstanding_coupon_form,
            'customer_pk': customer_pk,
            "url": reverse('create_customer_outstanding'),
            
            'page_title': 'Create customer supply',
            'customer_outstanding_page': True,
            'is_need_datetime_picker': True
        }
        
        return render(request,'client_management/customer_outstanding/create.html',context)


# customer count

def customer_count(request):
    routes = RouteMaster.objects.all()
    total_cash = 0
    total_credit = 0
    total_coupon = 0
    total_customers = 0
    customer_counts = []
    customer = Customers.objects.all().count()
    for route in routes:
        sales_man = ''
        van_route = Van_Routes.objects.filter(routes=route).first()
        if van_route and van_route.van:
            sales_man = van_route.van.salesman
        
        cash_count = Customers.objects.filter(routes=route, sales_type='CASH').count()
        credit_count = Customers.objects.filter(routes=route, sales_type='CREDIT').count()
        coupon_count = Customers.objects.filter(routes=route, sales_type__in=['CASH COUPON', 'CREDIT COUPON']).count()

        total_cash += cash_count
        total_credit += credit_count
        total_coupon += coupon_count
        total_customers += cash_count + credit_count + coupon_count
        if cash_count+credit_count+coupon_count != 0:
            customer_counts.append({
                'route_name': route.route_name,
                'sales_man': sales_man,
                'cash_count': cash_count,
                'credit_count': credit_count,
                'coupon_count': coupon_count,
                'total_customer': cash_count + credit_count + coupon_count
            })

    # customers with no route specified
    cash_count = Customers.objects.filter(routes=None, sales_type='CASH').count()
    credit_count = Customers.objects.filter(routes=None, sales_type='CREDIT').count()
    coupon_count = Customers.objects.filter(routes=None, sales_type__in=['CASH COUPON', 'CREDIT COUPON']).count()

    total_cash += cash_count
    total_credit += credit_count
    total_coupon += coupon_count
    total_customers += cash_count + credit_count + coupon_count
    
    customer_counts.append({
            'route_name': 'Not Specified',
            'sales_man': sales_man,
            'cash_count': cash_count,
            'credit_count': credit_count,
            'coupon_count': coupon_count,
            'total_customer': cash_count + credit_count + coupon_count
        })

    context = {
        'customer_counts': customer_counts,
        'total_cash': total_cash,
        'total_credit': total_credit,
        'total_coupon': total_coupon,
        'total_customers': total_customers,
    }
    # print('total customers:', total_customers)
    return render(request, 'client_management/customer_count.html', context)

def bottle_count(request):
    routes = RouteMaster.objects.all()

    context = {
        'instances': routes,
    }

    return render(request, 'client_management/bottle_count.html', context)
    

def bottle_count_route_wise(request, route_id):
    customers = Customers.objects.filter(routes__pk=route_id)
    
    context = {
        "instances" : customers
    }

    return render(request, 'client_management/route_details.html', context)
    
@login_required
def customer_orders(request):
    """
    Customer orders List
    :param request:
    :return: Customer orders list view
    """
    filter_data = {}
    instances = CustomerOrders.objects.all()
    
    query = request.GET.get("q")
    
    if query:

        instances = instances.filter(
            Q(customer__customer_name__icontains=query) |
            Q(customer__customer_id__icontains=query) |
            Q(customer__mobile_no__icontains=query) |
            Q(customer__whats_app__icontains=query) |
            Q(customer__email_id__icontains=query)
        )
        title = "Customer Order List - %s" % query
        filter_data['q'] = query
        
    acknowledge_form = CustomerOrdersAcknowledgeForm()
            
    context = {
        'instances': instances,
        'acknowledge_form': acknowledge_form,
        'page_name' : 'Customer Order List',
        'page_title' : 'Customer Order List',
        
        'is_customer_outstanding': True,
        'is_need_datetime_picker': True,
        'filter_data': filter_data,
    }

    return render(request, 'client_management/customer_order_list.html', context)

def customer_order_status_acknowledge(request,pk):
            
    try:
        with transaction.atomic():
            instance = CustomerOrders.objects.get(pk=pk)
            form = CustomerOrdersAcknowledgeForm(request.POST,instance=instance)

            data = form.save(commit=False)
            data.save()
            
            if data.order_status == "approve":
                DiffBottlesModel.objects.create(
                    product_item=data.product,
                    quantity_required=data.quantity,
                    delivery_date=data.delivery_date,
                    assign_this_to=data.customer.sales_staff,
                    mode="paid",
                    amount=data.total_amount,
                    discount_net_total=data.total_net_amount,
                    customer=data.customer,
                    created_by=data.created_by,
                    created_date=datetime.today(),
                )
                
            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Acknowledged",
                'reload': 'true',
            }
            
    except IntegrityError as e:
        # Handle database integrity error
        response_data = {
            "status": "false",
            "title": "Failed",
            "message": str(e),
        }

    except Exception as e:
        # Handle other exceptions
        response_data = {
            "status": "false",
            "title": "Failed",
            "message": str(e),
        }
    return HttpResponse(json.dumps(response_data), content_type='application/javascript')


def nonvisitreason_List(request):
    all_nonvisitreason= NonVisitReason.objects.all()
    context = {'all_nonvisitreason': all_nonvisitreason}
    return render(request, 'client_management/NonVisitReason/index_nonvisitReason.html', context)

def create_nonvisitreason(request):
    if request.method == 'POST':
        form = Create_NonVisitReasonForm(request.POST)
        if form.is_valid():
            data = form.save(commit=False)
            data.created_by = str(request.user.id)
            data.save()
            messages.success(request, 'Non Visit Reason created successfully!')
            return redirect('nonvisitreason_List')
        else:
            messages.error(request, 'Invalid form data. Please check the input.')
    else:
        form = Create_NonVisitReasonForm()
    context = {'form': form}
    return render(request, 'client_management/NonVisitReason/create_nonvisitreason.html', context)

def delete_nonvisitreason(request, id):
    delete_nonvisitreason = NonVisitReason.objects.get(id=id)
    if request.method == 'POST':
        delete_nonvisitreason.delete()
        return redirect('nonvisitreason_List')
    return render(request, 'client_management/NonVisitReason/delete_nonvisitreason.html', {'delete_nonvisitreason': delete_nonvisitreason})
