# views.py
import json
import random
from datetime import datetime, timedelta
from van_management.models import *
from django.db import IntegrityError
from django.http import HttpResponseRedirect, JsonResponse
from django.urls import reverse, reverse_lazy
from django.views import View
from django.shortcuts import render
from django.db.models import Q
# from client_management.models import CustodyCustomItems
from coupon_management.models import AssignStaffCouponDetails
from master.models import RouteMaster  # Assuming you have imported RouteMaster
from accounts.models import Customers
from product.models import Product, Product_Default_Price_Level
from sales_management.models import OutstandingLog
from tax_settings.models import Tax
from .forms import SaleEntryFilterForm
from django.views.generic import FormView, View
from django.http import JsonResponse
from django.urls import reverse_lazy
from .forms import CashCustomerSaleForm, CreditCustomerSaleForm, CashCouponCustomerSaleForm, CreditCouponCustomerSaleForm
from accounts.models import Customers
import random
import string
from django.db.models import Sum
from django.views.generic import View
from django.shortcuts import render, redirect
from django.db.models import Sum
from .models import  OutstandingLog, SaleEntryLog, SalesExtraModel, SalesTemp, Transaction
from accounts.models import Customers
from .forms import ProductForm  # Import the ProductForm we created earlier
from django.db.models import F, Value
from django.db.models.functions import Coalesce
from django.views.generic import ListView
from django.db.models import Count
from django.shortcuts import render, redirect, get_object_or_404
from io import BytesIO  

from .models import *
from customer_care.models import *
from client_management.models import *
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
import xlsxwriter
from .models import *
from openpyxl import Workbook
import pandas as pd
from io import BytesIO
from reportlab.pdfgen import canvas
from django.utils import timezone
from van_management.models import Van_Routes,Van,VanProductStock
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from invoice_management.models import *
from van_management.models import Expense

class TransactionHistoryListView(ListView):
    model = Transaction
    template_name = 'sales_management/transaction_history.html'  # Update with your actual template path
    context_object_name = 'transaction_history_list'

class SaleEntryLogListView(ListView):
    model = SaleEntryLog
    template_name = 'sales_management/sale_entry_history.html'
    context_object_name = 'sale_entry_logs'

class OutstandingLogListView(ListView):
    model = OutstandingLog
    template_name = 'sales_management/oustanding_history.html'  # Update with your actual template path
    context_object_name = 'outstanding_logs'

# views.py

from django.http import JsonResponse

def payment_submit(request):
    if request.method == 'POST' and request.is_ajax():
        # Retrieve form data
        total_amount = request.POST.get('total_amount')
        discount = request.POST.get('discount')
        net_taxable = request.POST.get('net_taxable')
        vat = request.POST.get('vat')
        total_to_collect = request.POST.get('total_to_collect')
        amount_received = request.POST.get('amount_received')
        balance = request.POST.get('balance')

        # Print or process the form data as needed
        print("Total Amount:", total_amount)
        print("Discount:", discount)
        print("Net Taxable:", net_taxable)
        print("VAT:", vat)
        print("Total to Collect:", total_to_collect)
        print("Amount Received:", amount_received)
        print("Balance:", balance)

        # Return a JSON response
        return JsonResponse({'message': 'Payment data received successfully'}, status=200)
    else:
        return JsonResponse({'error': 'Invalid request'}, status=400)


class SaleEntryLogView(View):
    template_name = 'sales_management/add_sale_entry.html'

    def get(self, request, *args, **kwargs):
        # Create an instance of the form and populate it with GET data
        form = SaleEntryFilterForm(request.GET)

        # Initialize not_found to False
        not_found = False

        # Check if the form is valid
        if form.is_valid():
            # Filter the queryset based on the form data
            route_filter = form.cleaned_data.get('route_name')
            search_query = form.cleaned_data.get('search_query')

            user_li = Customers.objects.all()

            if route_filter:
                user_li = user_li.filter(routes__route_name=route_filter)

            if search_query:
                # Apply search filter on relevant fields of the Customers model
                user_li = user_li.filter(
                    Q(customer_name__icontains=search_query) |
                    Q(building_name__icontains=search_query) |
                    Q(door_house_no__icontains=search_query)
                    # Add more fields as needed
                )

            # Check if the filtered data is empty
            not_found = not user_li.exists()

        else:
            # If the form is not valid, retrieve all customers
            user_li = Customers.objects.all()

        context = {'user_li': user_li, 'form': form, 'not_found': not_found}
        return render(request, self.template_name, context)
    
class CustomerDetailsView(View):
    template_name = 'sales_management/customer_sales_detail.html'

    def get(self, request, pk, *args, **kwargs):
        # Retrieve user details
        user_det = Customers.objects.get(customer_id=pk)
        sales_type = user_det.sales_type
        
        # Retrieve visit schedule data from user details
        visit_schedule_data = user_det.visit_schedule

        if visit_schedule_data:
            # Define a dictionary to map week names to numbers
            week_mapping = {"week1": 1, "week2": 2, "week3": 3, "week4": 4}

            # Initialize an empty list to store the result
            result = []

            # Loop through each day and its associated weeks
            for day, weeks in visit_schedule_data.items():
                for week in weeks:
                    # Extract week number using the week_mapping dictionary
                    week_number = week_mapping.get(week)
                    # Append day, week number, and day name to the result list
                    result.append((day, week_number))

            # Sort the result by week number
            result.sort(key=lambda x: x[1])

            # Prepare data for rendering
            data_for_rendering = []
            for slno, (day, week_number) in enumerate(result, start=1):
                data_for_rendering.append({'slno': slno, 'week': week_number, 'day': day})
        else:
            # If visit_schedule_data is None, handle it appropriately
            data_for_rendering = []

        # Filter AssignStaffCouponDetails based on customer_id
        assign_staff_coupon_details = AssignStaffCouponDetails.objects.filter(
            to_customer_id=user_det.customer_id
        )

        # Join AssignStaffCouponDetails with AssignStaffCoupon and aggregate the sum of remaining_quantity
        total_remaining_quantity = assign_staff_coupon_details.aggregate(
            total_remaining_quantity=Sum('staff_coupon_assign__remaining_quantity')
        )

        # Extract the sum of remaining_quantity from the aggregation result
        sum_remaining_quantity_coupons = total_remaining_quantity.get('total_remaining_quantity', 0)

        # Fetch all data from CustodyCustomItems model related to the user
        custody_items = CustodyCustomItems.objects.filter(customer=user_det)

        # Aggregate sum of coupons, empty bottles, and cash from OutstandingLog
        outstanding_log_aggregates = OutstandingLog.objects.filter(customer=user_det).aggregate(
            total_coupons=Sum('coupons'),
            total_empty_bottles=Sum('empty_bottles'),
            total_cash=Sum('cash')
        )
        # Check if all values in outstanding_log_aggregates are None
        if all(value is None for value in outstanding_log_aggregates.values()):
            outstanding_log_aggregates = None

        # Prepare the product form
        product_form = ProductForm()

        # Remove the coupon_method field from the form if sale type is "CASH" or "CREDIT"
        if sales_type in ["CASH", "CREDIT"]:
            del product_form.fields['coupon_method']
        # Add custody_items and aggregated data to the context
        context = {
            'user_det': user_det,
            'visit_schedule_data': data_for_rendering,
            'custody_items': custody_items,
            'outstanding_log_aggregates': outstanding_log_aggregates,  # Add aggregated data to the context
            'product_form': product_form,  # Add the product form to the context
        }

        return render(request, self.template_name, context)

    def post(self, request, pk, *args, **kwargs):
        # print("dddddddddddddddddddddddddddddddddddddddddddddddddddddddddddd")
        # Retrieve user details
        user_det = Customers.objects.get(customer_id=pk)

        # Process product form submission
        product_form = ProductForm(request.POST)
        if product_form.is_valid():
            # Save the product form data
            product = product_form.save(commit=False)
            product.created_by = request.user.username  # Set the created_by field
            product.save()
            return redirect('customer_details', pk=pk)
        else:
            # If the form is not valid, re-render the page with the form errors
            context = {
                'user_det': user_det,
                'product_form': product_form,
                # Add other context data as needed
            }
            return render(request, self.template_name, context)

class GetProductsByCategoryView(View):
    def get(self, request):
        category_id = request.GET.get('category_id')
        print(category_id,'category_id')
        products = Product.objects.filter(category_id=category_id).values('product_id', 'product_name')
        print("productsproducts", products)
        return JsonResponse({'products': list(products)})
    


    

class InitiateSaleView(FormView, View):
    template_name = 'sales_management/customer_sale_form.html'
    success_url = reverse_lazy('sale_entry_log_view')

    def get_form_class(self, sales_type):
        if sales_type == 'CASH':
            return CashCustomerSaleForm
        elif sales_type == 'CREDIT':
            return CreditCustomerSaleForm
        elif sales_type == 'CASH COUPON':
            return CashCouponCustomerSaleForm
        elif sales_type == 'CREDIT COUPON':
            return CreditCouponCustomerSaleForm
        else:
            return None

    def get(self, request, *args, **kwargs):
        print("///////////////////////////////////////////////////////////////////////////")
        get_data = request.GET.dict()
        print("get_dataget_dataget_data", get_data)
        customer_id = request.GET.get('customer_id')
        product_id = request.GET.get('product_name')
        print("product_idproduct_id", product_id)
        coupon_method = request.GET.get('coupon_method')
        print("coupon_methodcoupon_method", coupon_method)
        # Fetch the customer object
        customer = Customers.objects.filter(customer_id=customer_id).first()
        # Check if the customer exists
        # Filter AssignStaffCouponDetails based on customer_id
        assign_staff_coupon_details = AssignStaffCouponDetails.objects.filter(
            to_customer_id=customer_id
        )

        # Join AssignStaffCouponDetails with AssignStaffCoupon and aggregate the sum of remaining_quantity
        total_remaining_quantity = assign_staff_coupon_details.aggregate(
            total_remaining_quantity=Sum('staff_coupon_assign__remaining_quantity')
        )

        # Extract the sum of remaining_quantity from the aggregation result
        sum_remaining_quantity_coupons = total_remaining_quantity.get('total_remaining_quantity', 0)

        print("sum_remaining_quantity_couponssum_remaining_quantity_coupons", sum_remaining_quantity_coupons)
        if customer:
            if coupon_method != 'digital' or customer.sales_type not in ['CREDIT COUPON', 'CASH COUPON']:
                # Get the sales type for the customer
                sales_type = customer.sales_type
                # Instantiate the appropriate form class based on sales type
                form_class = self.get_form_class(sales_type)
                if form_class:
                    form = form_class()
                    # test 
                else:
                    return JsonResponse({'error': 'Invalid sales type'}, status=400)

                return self.render_to_response({'sum_remaining_quantity_coupons': sum_remaining_quantity_coupons, 'sales_type': sales_type, 'customer_id': customer_id, 'product_id': product_id, 'form': form, 'coupon_method': coupon_method})
                
            # Return 404 if customer does not exist
            return JsonResponse({'error': 'digital coupon found'}, status=404)
                            
        # Return 404 if customer does not exist
        return JsonResponse({'error': 'Customer not found'}, status=404)



    def generate_invoice_number(self):
        # Generate a random 10-digit invoice number
        invoice_number = ''.join(random.choices(string.digits, k=10))
        
        # Check if the generated number is unique in Transaction table
        while Transaction.objects.filter(invoice_number=invoice_number).exists():
            # If not unique, generate a new invoice number
            invoice_number = ''.join(random.choices(string.digits, k=10))
        
        # Check if the generated number is unique in SalesTemp table
        while SalesTemp.objects.filter(invoice_number=invoice_number).exists():
            # If not unique, generate a new invoice number
            invoice_number = ''.join(random.choices(string.digits, k=10))
        
        return invoice_number
    
    def delete_sales_temp_data(self,invoice_number):
        try:
            sales_temp_entry = SalesTemp.objects.get(invoice_number=invoice_number)
            sales_temp_entry.delete()
            print("Data deleted successfully from SalesTemp table for invoice number:", invoice_number)
        except SalesTemp.DoesNotExist:
            print("No data found in SalesTemp table for invoice number:", invoice_number)

    
    def post(self, request, *args, **kwargs):
        sales_type = request.POST.get('sales_type')
        customer_id = request.POST.get('customer_id')
        product_id = request.POST.get('product_id')
        customer = Customers.objects.filter(customer_id=customer_id).first()
        product = Product.objects.filter(product_id=product_id).first()
        id_status = request.POST.get('status')
        # Get sales_type from form data
        print("qqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqq", request.POST)
        print("id_statusid_statusid_status", id_status)
        # Instantiate the appropriate form class based on sales type
        form_class = self.get_form_class(sales_type)
        print("form_classform_class", form_class)
        if form_class:
            print("entry_________________!")
            form = form_class(request.POST)
            print("formform", form)
            # Convert form input values to integers
            qty_needed = int(request.POST.get('qty_needed', 0))
            no_of_coupons = int(request.POST.get('no_of_coupons', 0))
            coupon_variations = int(request.POST.get('coupon_variations', 0))
            print("coupon_variationscoupon_variations", coupon_variations)
            empty_bottles = int(request.POST.get('empty_bottles', 0))
            collected_bottles = int(request.POST.get('collected_bottles', 0))
            bottle_variations = int(request.POST.get('bottle_variations', 0))
            id_status = request.POST.get('status')
            invoice_number = self.generate_invoice_number()
            cash=0
            print("id_statusid_status", id_status)
            if id_status == 'PAID':
                print("eeeeeeeeeeeeeeeeeeeeeeeee")
                invoice_number=request.POST.get('invoice_number')
                sales_temp_data = SalesTemp.objects.get(invoice_number=invoice_number).data
                customer = Customers.objects.filter(customer_id=sales_temp_data['customer_id']).first()
                product = Product.objects.filter(product_id=sales_temp_data['product_id']).first()
                get_tax_percentage= Tax.objects.filter(name="vat").first()
                vat_value= get_tax_percentage/100
                print("__________________________________________________________________________________", invoice_number)
                print("sales_temp_datasales_temp_data", sales_temp_data)
                print("__________________________________________________________________________________")
                # Calculate data based on form fields
                total_amount = (qty_needed + no_of_coupons) * coupon_variations  # Dummy calculation for total_amount
                discount = total_amount * 0.1  # Dummy calculation for discount (10% of total_amount)
                net_taxable = total_amount - discount  # Dummy calculation for net_taxable
                vat = net_taxable * vat_value  # Dummy calculation for VAT (5% of net_taxable)
                received_amount = net_taxable - vat  # Dummy calculation for received_amount (net_taxable - VAT)
                balance = received_amount - total_amount  # Dummy calculation for balance (received_amount - total_amount)
                print("id_statusid_status", id_status)
                cash=total_amount
                # Generate initial invoice number
                try:
                    # Save data to Transaction model
                    transaction = Transaction.objects.create(
                        customer=customer,
                        category='INCOME',
                        amount=total_amount,
                        created_staff=request.user,
                        invoice_number=invoice_number,
                        transaction_category=id_status
                        # Add other fields as needed
                    )
                    # Save data to SaleEntryLog model
                    sale_entry = SaleEntryLog.objects.create(
                        customer=customer,
                        quantity=qty_needed,
                        total_amount=total_amount,
                        discount=discount,
                        net_taxable=net_taxable,
                        vat=vat,
                        received_amount=received_amount,
                        balance=balance,
                        product=product,
                        empty_bottles=empty_bottles,

                        # Add other fields as needed
                    )
                    # Create SalesExtraModel instance with the calculated values
                    sales_extra_model = SalesExtraModel.objects.create(
                        qty_needed=qty_needed,
                        no_of_coupons=no_of_coupons,
                        coupon_variations=coupon_variations,
                        empty_bottles=empty_bottles,
                        collected_bottles=collected_bottles,
                        bottle_variations=bottle_variations,
                        status=id_status,
                        order_number=sale_entry.order_number  # Assuming you want to link it with the same order number
                    )
                    # Create an entry in the OutstandingLog model
                    outstanding_log = OutstandingLog.objects.create(
                        customer_id=customer_id,
                        coupons=coupon_variations,
                        empty_bottles=bottle_variations,
                        cash=balance,
                        created_by=request.user  # Assuming you have a logged-in user
                    )

                    self.delete_sales_temp_data(invoice_number)

                    return HttpResponseRedirect(reverse('sale_entry_log_list'))

                except IntegrityError as e:
                    # Extract relevant information from the IntegrityError object
                    error_message = str(e)
                    # Return a JsonResponse with the error message and appropriate status code
                    return JsonResponse({'error': error_message}, status=400)

            if id_status == "FOC":
                print("FOC is working")
                try:
                    # Save data to Transaction model
                    transaction = Transaction.objects.create(
                        customer=customer,
                        category='EXPENSE',
                        amount=0,
                        created_staff=request.user,
                        invoice_number=invoice_number,
                        transaction_category=id_status
                        # Add other fields as needed
                    )
                    # Save data to SaleEntryLog model
                    sale_entry = SaleEntryLog.objects.create(
                        customer=customer,
                        quantity=qty_needed,
                        total_amount=0,
                        discount=0,
                        net_taxable=0,
                        vat=0,
                        received_amount=0,
                        balance=0,
                        product=product,
                        empty_bottles=empty_bottles,
                        # Add other fields as needed
                    )

                    # Create SalesExtraModel instance with the calculated values
                    sales_extra_model = SalesExtraModel.objects.create(
                        qty_needed=qty_needed,
                        no_of_coupons=no_of_coupons,
                        coupon_variations=coupon_variations,
                        empty_bottles=empty_bottles,
                        collected_bottles=collected_bottles,
                        bottle_variations=bottle_variations,
                        status=id_status,
                        order_number=sale_entry.order_number  # Assuming you want to link it with the same order number
                    )
                    outstanding_log = OutstandingLog.objects.create(
                        customer_id=customer_id,
                        coupons=coupon_variations,
                        empty_bottles=bottle_variations,
                        cash=0,
                        created_by=request.user  # Assuming you have a logged-in user
                    )

                    return HttpResponseRedirect(reverse('sale_entry_log_list'))

                except IntegrityError as e:
                    # Extract relevant information from the IntegrityError object
                    error_message = str(e)
                    # Return a JsonResponse with the error message and appropriate status code
                    return JsonResponse({'error': error_message}, status=400)
                
            if id_status == "PENDING":
                print("PENDING is working")
                try:
                    # Save data to Transaction model
                    transaction = Transaction.objects.create(
                        customer=customer,
                        category='INCOME',
                        amount=0,
                        created_staff=request.user,
                        invoice_number=invoice_number,
                        transaction_category=id_status
                        # Add other fields as needed
                    )
                    # Save data to SaleEntryLog model
                    sale_entry = SaleEntryLog.objects.create(
                        customer=customer,
                        quantity=qty_needed,
                        total_amount=0,
                        discount=0,
                        net_taxable=0,
                        vat=0,
                        received_amount=0,
                        balance=0,
                        product=product,
                        empty_bottles=empty_bottles,
                        # Add other fields as needed
                    )
                    return HttpResponseRedirect(reverse('sale_entry_log_list'))

                except IntegrityError as e:
                    # Extract relevant information from the IntegrityError object
                    error_message = str(e)
                    # Return a JsonResponse with the error message and appropriate status code
                    return JsonResponse({'error': error_message}, status=400)
                
                
            if id_status == "CUSTODY":
                print("CUSTODY is working")
                try:
                    # Save data to Transaction model
                    transaction = Transaction.objects.create(
                        customer=customer,
                        category='INCOME',
                        amount=0,
                        created_staff=request.user,
                        invoice_number=invoice_number,
                        transaction_category=id_status
                        # Add other fields as needed
                    )
                    # Save data to SaleEntryLog model
                    sale_entry = SaleEntryLog.objects.create(
                        customer=customer,
                        quantity=qty_needed,
                        total_amount=0,
                        discount=0,
                        net_taxable=0,
                        vat=0,
                        received_amount=0,
                        balance=0,
                        product=product,
                        empty_bottles=empty_bottles,
                        # Add other fields as needed
                    )
                    outstanding_log = OutstandingLog.objects.create(
                        customer_id=customer_id,
                        coupons=coupon_variations,
                        empty_bottles=bottle_variations,
                        cash=0,
                        created_by=request.user  # Assuming you have a logged-in user
                    )
                    return HttpResponseRedirect(reverse('sale_entry_log_list'))

                except IntegrityError as e:
                    # Extract relevant information from the IntegrityError object
                    error_message = str(e)
                    # Return a JsonResponse with the error message and appropriate status code
                    return JsonResponse({'error': error_message}, status=400)
            # else:
            #     # Print form name
            #     form_name = form.__class__.__name__  # Get the class name of the form
            #     print(f"{form_name} is not valid. Errors:")
            #     for field, errors in form.errors.items():
            #         print(f"- Field: {field}, Errors: {', '.join(errors)}")
            #     # Return a JsonResponse with the error message and appropriate status code
            #     return JsonResponse({'error': f'{form_name} is not valid'}, status=400)


        # Ensure that you return a response in all code paths
        return JsonResponse({'success': 'Sale entry created successfully'}, status=200)



from django.http import JsonResponse

class PaymentForm(View):
    template_name = 'sales_management/customer_sales_detail.html'

    def generate_invoice_number(self):
        # Generate a random 10-digit invoice number
        invoice_number = ''.join(random.choices(string.digits, k=10))
        # Check if the generated number is unique in Transaction table
        while Transaction.objects.filter(invoice_number=invoice_number).exists():
            # If not unique, generate a new invoice number
            invoice_number = ''.join(random.choices(string.digits, k=10))
        # Check if the generated number is unique in SalesTemp table
        while SalesTemp.objects.filter(invoice_number=invoice_number).exists():
            # If not unique, generate a new invoice number
            invoice_number = ''.join(random.choices(string.digits, k=10))        
        return invoice_number

    def post(self, request, *args, **kwargs):
        print("gggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggg", request.POST)
        # Retrieve data from the POST request
        customer_id = request.POST.get('customer_id')
        product_id = request.POST.get('product_id')
        sales_type = request.POST.get('sales_type')
        qty_needed = request.POST.get('qty_needed')
        empty_bottles = request.POST.get('empty_bottles')
        collected_bottles = request.POST.get('collected_bottles')
        status = request.POST.get('status')
        discount_percentage = float(request.POST.get('discount_percentage'))/100
        print('statusstatusstatusstatusstatus', status)
        invoice_number = self.generate_invoice_number()
        if status == "PAID":
            # Create a dictionary to store the data
            data = {
                'customer_id': customer_id,
                'product_id': product_id,
                'sales_type': sales_type,
                'qty_needed': qty_needed,
                'empty_bottles': empty_bottles,
                'collected_bottles': collected_bottles,
                'status': status,
                'discount_percentage': discount_percentage,
            }
            # Create an instance of SalesTemp model and save it to the database
            sales_temp = SalesTemp.objects.create(invoice_number=invoice_number, data=data)

        # Fetch the rate from the Customers table based on the customer_id
        customer = Customers.objects.filter(customer_id=customer_id).first()
        if customer:
            customer_type = customer.customer_type
            get_customer_default_price = Product_Default_Price_Level.objects.filter(customer_type=customer_type, product_id=product_id).first()
            if get_customer_default_price:
                rate = float(get_customer_default_price.rate)
            else:
                # Handle the case where the rate is not available
                return JsonResponse({'error': 'Rate not available for the customer and product'}, status=400)
            # Fetch tax options from the Tax model
            tax_options = Tax.objects.all().values_list('id', 'name')
            # Calculate the total_amount by multiplying rate with qty_needed
            # Add None at the beginning of the tax options list
            tax_options_list = list(tax_options)
            tax_options_list.insert(0, (None, 'None'))
            
            total_amount = float(qty_needed) * rate

            print("7//////////////////////////////////////////////////////")
            print("qty_needed", qty_needed, "rateraterate", rate)
            print("total_amounttotal_amount", total_amount, type(total_amount))
            print("discount_percentage", discount_percentage, type(discount_percentage))


            discount_amount = total_amount * discount_percentage
            net_taxable = total_amount - discount_amount
            # Calculate VAT (5% of total_amount
            # Query to retrieve product tax percentage
            product_tax_percentage = Product.objects.filter(product_id=product_id).annotate(
                tax_percentage=Coalesce(F('tax__percentage'), Value(0))
            ).values_list('tax_percentage', flat=True).first()

            vat_percentage=product_tax_percentage/100
            print("vat_percentage", vat_percentage)
            vat = total_amount * vat_percentage

            total_to_collect = total_amount + vat
            # Calculate total_to_collect (net_taxable plus VAT)
        else:
            # Handle the case where customer is not found
            return JsonResponse({'error': 'Customer not found'}, status=400)

        # Prepare the response data
        response_data = {
            'total_amount': total_amount,
            'discount_amount': discount_amount,
            'net_taxable':net_taxable,
            'total_to_collect': total_to_collect,
            'vat':vat,
            'invoice_number':invoice_number,
            'sales_type': sales_type,
            'status': status,
        }
        print("response_dataresponse_data", response_data)

        # Return the response
        return JsonResponse(response_data)


class CalculateTotaltoCollect(View):
    def get(self, request):
        print("colllectAmount__calc")
        vat_value = request.GET.get('vat_value')
        # products = Product.objects.filter(vat_value=vat_value).values('product_id', 'product_name')
        # print("productsproducts", products)
        return JsonResponse({'products': list(vat_value)})



# class CouponSaleView(View):
#     template_name = 'sales_management/coupon_sale.html'

#     def get(self, request, *args, **kwargs):
#         # Create an instance of the form and populate it with GET data
#         form = SaleEntryFilterForm(request.GET)

#         # Initialize not_found to False
#         not_found = False

#         # Check if the form is valid
#         if form.is_valid():
#             # Filter the queryset based on the form data
#             route_filter = form.cleaned_data.get('route_name')
#             search_query = form.cleaned_data.get('search_query')

#             user_li = Customers.objects.all()

#             if route_filter:
#                 user_li = user_li.filter(routes__route_name=route_filter)

#             if search_query:
#                 # Apply search filter on relevant fields of the Customers model
#                 user_li = user_li.filter(
#                     Q(customer_name__icontains=search_query) |
#                     Q(building_name__icontains=search_query) |
#                     Q(door_house_no__icontains=search_query)
#                     # Add more fields as needed
#                 )

#             not_found = not user_li.exists()

#         else:
#             # If the form is not valid, retrieve all customers
#             user_li = Customers.objects.all()

#         context = {'user_li': user_li, 'form': form, 'not_found': not_found}
#         return render(request, self.template_name, context)



#     def post(self, request, *args, **kwargs):
        
    
class CouponSaleView(View):
    template_name = 'sales_management/coupon_sale.html'

    def get(self, request, *args, **kwargs):
        # Create an instance of the form and populate it with GET data
        form = SaleEntryFilterForm(request.GET)

        # Initialize not_found to False
        not_found = False

        # Check if the form is valid
        if form.is_valid():
            # Filter the queryset based on the form data
            route_filter = form.cleaned_data.get('route_name')
            search_query = form.cleaned_data.get('search_query')

            user_li = Customers.objects.all()

            if route_filter:
                user_li = user_li.filter(routes__route_name=route_filter)

            if search_query:
                user_li = user_li.filter(
                    Q(customer_name__icontains=search_query) |
                    Q(building_name__icontains=search_query) |
                    Q(door_house_no__icontains=search_query)
                )

            not_found = not user_li.exists()

        else:
            user_li = CustomerCoupons.objects.filter()

        balance_coupons = user_li
        print(balance_coupons,'balance_coupons')

        # Get the manual book type last purchased
        # manual_book_type_last_purchased = user_li.latest('created_date').deposit_type

        context = {'user_li': user_li, 'form': form, 'not_found': not_found}
        return render(request, self.template_name, context)
    

class DetailsView(View):
    template_name = 'sales_management/detail.html'

    def get(self, request, pk, *args, **kwargs):
        # Retrieve user details
        user_det = Customers.objects.get(customer_id=pk)
        sales_type = user_det.sales_type
        
        # Retrieve visit schedule data from user details
        visit_schedule_data = user_det.visit_schedule

        if visit_schedule_data:
            # Define a dictionary to map week names to numbers
            week_mapping = {"week1": 1, "week2": 2, "week3": 3, "week4": 4}

            # Initialize an empty list to store the result
            result = []

            # Loop through each day and its associated weeks
            for day, weeks in visit_schedule_data.items():
                for week in weeks:
                    # Extract week number using the week_mapping dictionary
                    week_number = week_mapping.get(week)
                    # Append day, week number, and day name to the result list
                    result.append((day, week_number))

            # Sort the result by week number
            # result.sort(key=lambda x: x[1])

            # Prepare data for rendering
            data_for_rendering = []
            for slno, (day, week_number) in enumerate(result, start=1):
                data_for_rendering.append({'slno': slno, 'week': week_number, 'day': day})
        else:
            # If visit_schedule_data is None, handle it appropriately
            data_for_rendering = []

        # Filter AssignStaffCouponDetails based on customer_id
        assign_staff_coupon_details = AssignStaffCouponDetails.objects.filter(
            to_customer_id=user_det.customer_id
        )

        # Join AssignStaffCouponDetails with AssignStaffCoupon and aggregate the sum of remaining_quantity
        total_remaining_quantity = assign_staff_coupon_details.aggregate(
            total_remaining_quantity=Sum('staff_coupon_assign__remaining_quantity')
        )

        # Extract the sum of remaining_quantity from the aggregation result
        sum_remaining_quantity_coupons = total_remaining_quantity.get('total_remaining_quantity', 0)

        # Fetch all data from CustodyCustomItems model related to the user
        custody_items = CustodyCustomItems.objects.filter(customer=user_det)

        # Aggregate sum of coupons, empty bottles, and cash from OutstandingLog
        outstanding_log_aggregates = OutstandingLog.objects.filter(customer=user_det).aggregate(
            total_coupons=Sum('coupons'),
            total_empty_bottles=Sum('empty_bottles'),
            total_cash=Sum('cash')
        )
        # Check if all values in outstanding_log_aggregates are None
        if all(value is None for value in outstanding_log_aggregates.values()):
            outstanding_log_aggregates = None

        # Prepare the product form
        product_form = ProductForm()

        # Remove the coupon_method field from the form if sale type is "CASH" or "CREDIT"
        if sales_type in ["CASH", "CREDIT"]:
            del product_form.fields['coupon_method']
        # Add custody_items and aggregated data to the context
        context = {
            'user_det': user_det,
            'visit_schedule_data': data_for_rendering,
            'custody_items': custody_items,
            'outstanding_log_aggregates': outstanding_log_aggregates,  # Add aggregated data to the context
            'product_form': product_form,  # Add the product form to the context
        }

        return render(request, self.template_name, context)

    def post(self, request, pk, *args, **kwargs):
        print("dddddddddddddddddddddddddddddddddddddddddddddddddddddddddddd")
        # Retrieve user details
        user_det = Customers.objects.get(customer_id=pk)

        # Process product form submission
        product_form = ProductForm(request.POST)
        if product_form.is_valid():
            # Save the product form data
            product = product_form.save(commit=False)
            product.created_by = request.user.username  # Set the created_by field
            product.save()
            return redirect('customer_details', pk=pk)
        else:
            # If the form is not valid, re-render the page with the form errors
            context = {
                'user_det': user_det,
                'product_form': product_form,
                # Add other context data as needed
            }
            return render(request, self.template_name, context)
        
#------------------SALES REPORT-------------------------        

        
def salesreport(request):
    instances = CustomUser.objects.filter(user_type='Salesman')
    print("instances",instances)
    return render(request, 'sales_management/sales_report.html', {'instances': instances})

def salesreportview(request, salesman):
    salesman = get_object_or_404(CustomUser, pk=salesman)
    customer_supplies = CustomerSupply.objects.filter(salesman=salesman)
    customer_supply_items = CustomerSupplyItems.objects.filter(customer_supply__in=customer_supplies)

    # Calculate counts of products for each sales type
    cash_coupon_counts = customer_supplies.filter(customer__sales_type='CASH COUPON').aggregate(total_products=Count('customersupplyitems'))

    # print("cash_coupon_counts",cash_coupon_counts)
    credit_coupon_counts = customer_supplies.filter(customer__sales_type='CREDIT COUPON').aggregate(total_products=Count('customersupplyitems'))
    # print("credit_coupon_counts",credit_coupon_counts)
    cash_counts = customer_supplies.filter(customer__sales_type='CASH').aggregate(total_products=Count('customersupplyitems'))
    # print("cash_counts",cash_counts)
    credit_counts = customer_supplies.filter(customer__sales_type='CREDIT').aggregate(total_products=Count('customersupplyitems'))
    # print("credit_counts",credit_counts)
    
    context = {
        'salesman': salesman,
        'customer_supplies':customer_supplies,
        'customer_supply_items':customer_supply_items,
        'cash_coupon_counts': cash_coupon_counts['total_products'],
        'credit_coupon_counts': credit_coupon_counts['total_products'],
        'cash_counts': cash_counts['total_products'],
        'credit_counts': credit_counts['total_products'],
    }

    return render(request, 'sales_management/salesreportview.html', context)

def download_salesreport_pdf(request):
    # Retrieve sales report data
    customer_supplies = CustomerSupply.objects.all()
    # Check if customer_supplies is not empty
    if customer_supplies:
        # Create the HTTP response with PDF content type and attachment filename
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'

        # Create a PDF document
        pdf_buffer = BytesIO()
        pdf = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        data = []

        # Add headers
        headers = ['Customer Name', 'Customer Address', 'Customer Type', 'Customer Sales Type', 
                   'Cash Coupon Quantity','Credit Coupon Quantity','Cash Quantity','Credit Quantity']
        data.append(headers)


        # Add data to the PDF document
        sl_no = 1
        for supply in customer_supplies:
            try:
                cash_coupon_counts = supply.customersupplyitems_set.filter(customer_supply=supply, customer_supply__customer__sales_type='CASH COUPON').count()
                credit_coupon_counts = supply.customersupplyitems_set.filter(customer_supply=supply, customer_supply__customer__sales_type='CREDIT COUPON').count()
                cash_counts = supply.customersupplyitems_set.filter(customer_supply=supply, customer_supply__customer__sales_type='CASH').count()
                credit_counts = supply.customersupplyitems_set.filter(customer_supply=supply, customer_supply__customer__sales_type='CREDIT').count()
            except Exception as e:
                print(e)
                cash_coupon_counts = 0
                credit_coupon_counts = 0
                cash_counts = 0
                credit_counts = 0
            
            # Append data for each supply
            data.append([
                sl_no,
                supply.customer.customer_name,
                f"{supply.customer.building_name} {supply.customer.door_house_no}",
                supply.customer.customer_type,
                supply.customer.sales_type,
                cash_coupon_counts,
                credit_coupon_counts,
                cash_counts,
                credit_counts
            ])
            sl_no += 1

        table = Table(data)
        style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)])
        table.setStyle(style)

        # Add the table to the PDF document
        elements = [table]
        pdf.build(elements)

        # Get the value of the BytesIO buffer and write it to the response
        pdf_value = pdf_buffer.getvalue()
        pdf_buffer.close()
        response.write(pdf_value)

        return response
    else:
        return HttpResponse("No data available for sales report.")
    

def download_salesreport_excel(request):
    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active

    # Add headers to the worksheet
    headers = ['Sl No', 'Customer Name', 'Customer Address', 'Customer Type', 
               'Customer Sales Type', 'Cash Coupon Quantity', 'Credit Coupon Quantity', 
               'Cash Quantity', 'Credit Quantity']
    worksheet.append(headers)

    # Retrieve sales report data
    customer_supplies = CustomerSupply.objects.all()

    # Add data to the worksheet
    sl_no = 1
    for supply in customer_supplies:
        try:
            
            cash_coupon_counts = supply.customersupplyitems_set.filter(customer_supply=supply, customer_supply__customer__sales_type='CASH COUPON').count()
            print("cash_coupon_counts",cash_coupon_counts)
            credit_coupon_counts = supply.customersupplyitems_set.filter(customer_supply=supply, customer_supply__customer__sales_type='CREDIT COUPON').count()
            print("credit_coupon_counts",credit_coupon_counts)
            cash_counts = supply.customersupplyitems_set.filter(customer_supply=supply, customer_supply__customer__sales_type='CASH').count()
            print("cash_counts",cash_counts)
            credit_counts = supply.customersupplyitems_set.filter(customer_supply=supply, customer_supply__customer__sales_type='CREDIT').count()
            print("credit_counts",credit_counts)

        except Exception as e:
            print(e)
            cash_coupon_counts = 0
            credit_coupon_counts = 0
            cash_counts = 0
            credit_counts = 0
        
        # Write data for each supply to the worksheet
        row_data = [
            sl_no,
            supply.customer.customer_name,
            f"{supply.customer.building_name} {supply.customer.door_house_no}",
            supply.customer.customer_type,
            supply.customer.sales_type,
            cash_coupon_counts,
            credit_coupon_counts,
            cash_counts,
            credit_counts
        ]
        worksheet.append(row_data)

        sl_no += 1

    # Create HTTP response with Excel content type and attachment filename
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'

    # Write the workbook data into the response
    workbook.save(response)

    # Write headers
    headers = ['Customer Name', 'Customer Address', 'Product', 'Quantity', 'Amount']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)

    # Write data to the Excel workbook
    row = 1  # Start from row 1 since row 0 is for headers
    for supply in customer_supplies:
        for item in supply.customersupplyitems_set.all():
            customer_address = ", ".join(filter(None, [supply.customer.building_name, supply.customer.door_house_no, supply.customer.floor_no]))
            product_name = "Product Not Found"
            if item.product:  # Check if product is not None
                product_name = item.product.product_name
            worksheet.write(row, 0, supply.customer.customer_name)
            worksheet.write(row, 1, customer_address)
            worksheet.write(row, 2, product_name)
            worksheet.write(row, 3, item.quantity)
            worksheet.write(row, 4, item.amount)
            row += 1

    workbook.close()

    return response

#------------------Collection Report-------------------------                
# def collectionreport(request):
#     start_date = None
#     end_date = None
#     selected_date = None
#     selected_route_id = None
#     selected_route = None
#     template = 'sales_management/collection_report.html'
#     colectionpayment = CollectionPayment.objects.all()
    
#     routes = RouteMaster.objects.all()
#     route_counts = {}
#     today = datetime.today()
    
#     if request.method == 'POST':
#         start_date = request.POST.get('start_date')
#         end_date = request.POST.get('end_date')
#         selected_date = request.POST.get('date')
#         selected_route_id = request.POST.get('selected_route_id')
#         if start_date and end_date:
#             colectionpayment = colectionpayment.filter(customer_supply__created_date__range=[start_date, end_date])
#         elif selected_date:
#             colectionpayment = colectionpayment.filter(customer_supply__created_date=selected_date)
        
#         if selected_route_id:
#             selected_route = RouteMaster.objects.get(id=selected_route_id)
#             colectionpayment = colectionpayment.filter(customer__routes__route_name=selected_route)
    
#     # /
    
#     context = {
#         'colectionpayment': colectionpayment, 
#         'routes': routes, 
#         'route_counts': route_counts, 
#         'today': today,
#         'start_date': start_date, 
#         'end_date': end_date, 
#         'selected_date': selected_date, 
#         'selected_route_id': selected_route_id, 
#         'selected_route': selected_route,
        
#     }
#     return render(request, template, context)



# def collection_report_excel(request):
#     instances = CollectionPayment.objects.all()
#     route_filter = request.GET.get('route_name')
#     start_date_str = request.GET.get('start_date')
#     end_date_str = request.GET.get('end_date')
    
#     if start_date_str and end_date_str:
#         start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
#         end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
#         instances = instances.filter(customer__customer_supply__created_date__range=[start_date, end_date])
    
#     print('route_filter :', route_filter)
#     if route_filter and route_filter != '' and route_filter != 'None':
#         instances = instances.filter(routes__route_name=route_filter)

#     route_li = RouteMaster.objects.all()
#     serial_number = 1
#     for customer in instances:
#         customer.serial_number = serial_number
#         serial_number += 1
#     data = {
#         'Serial Number': [customer.serial_number for customer in instances],
#         'Date': [customer.customer_supply.created_date.date() for customer in instances],
#         'Customer name': [customer.customer.customer_name for customer in instances],
#         'Mobile No': [customer.customer.mobile_no for customer in instances],
#         'Route': [customer.customer.routes.route_name if customer.customer.routes else '' for customer in instances],
#         'Building Name': [customer.customer.building_name for customer in instances],
#         'House No': [customer.customer.door_house_no if customer.customer.door_house_no else 'Nil' for customer in instances],
#         'Receipt No/Reference No': [customer.customer_supply.reference_number for customer in instances],
#         'Amount': [customer.amount for customer in instances],
#         'Mode of Payment': [customer.payment_method for customer in instances],

#     }
#     df = pd.DataFrame(data)

#     buffer = BytesIO()
#     with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
#         df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=4)
#         workbook = writer.book
#         worksheet = writer.sheets['Sheet1']
#         table_border_format = workbook.add_format({'border':1})
#         worksheet.conditional_format(4, 0, len(df.index)+4, len(df.columns) - 1, {'type':'cell', 'criteria': '>', 'value':0, 'format':table_border_format})
#         merge_format = workbook.add_format({'align': 'center', 'bold': True, 'font_size': 16, 'border': 1})
#         worksheet.merge_range('A1:J2', f'National Water', merge_format)
#         merge_format = workbook.add_format({'align': 'center', 'bold': True, 'border': 1})
#         worksheet.merge_range('A3:J3', f'    Collection Report   ', merge_format)
#         merge_format = workbook.add_format({'align': 'center', 'bold': True, 'border': 1})
#         worksheet.merge_range('A4:J4', '', merge_format)
    
#     filename = f"Collection Report.xlsx"
#     response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'inline; filename = "{filename}"'
#     return response


# def dailycollectionreport(request):
#     instances = CollectionPayment.objects.all()
#     route_filter = request.GET.get('route_name')
#     start_date_str = request.GET.get('start_date')

#     if start_date_str :
#         start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
#         instances = instances.filter(created_date__range=start_date)
#     if route_filter:
#             instances = instances.filter(routes__route_name=route_filter)
#     route_li = RouteMaster.objects.all()
    
#     context = {'instances': instances,'route_li':route_li}
#     return render(request, 'sales_management/daily_collection_report.html', context)


# def daily_collection_report_excel(request):
#     instances = CollectionPayment.objects.all()
#     route_filter = request.GET.get('route_name')
#     start_date_str = request.GET.get('start_date')
    
#     if start_date_str :
#         start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
#         instances = instances.filter(customer__customer_supply__created_date__range=start_date)
    
#     print('route_filter :', route_filter)
#     if route_filter and route_filter != '' and route_filter != 'None':
#         instances = instances.filter(routes__route_name=route_filter)

#     route_li = RouteMaster.objects.all()
#     serial_number = 1
#     for customer in instances:
#         customer.serial_number = serial_number
#         serial_number += 1
#     data = {
#         'Serial Number': [customer.serial_number for customer in instances],
#         'Customer name': [customer.customer.customer_name for customer in instances],
#         'Mobile No': [customer.customer.mobile_no for customer in instances],
#         'Route': [customer.customer.routes.route_name if customer.customer.routes else '' for customer in instances],
#         'Building Name': [customer.customer.building_name for customer in instances],
#         'House No': [customer.customer.door_house_no if customer.customer.door_house_no else 'Nil' for customer in instances],
#         'Receipt No/Reference No': [customer.customer_supply.reference_number for customer in instances],
#         'Amount': [customer.amount for customer in instances],
#         'Mode of Payment': [customer.payment_method for customer in instances],
#         'Invoice': [customer.invoice.invoice_no for customer in instances],
#         'Invoice Reference No': [customer.invoice.reference_no  for customer in instances],


#     }
#     df = pd.DataFrame(data)

#     buffer = BytesIO()
#     with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
#         df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=4)
#         workbook = writer.book
#         worksheet = writer.sheets['Sheet1']
#         table_border_format = workbook.add_format({'border':1})
#         worksheet.conditional_format(4, 0, len(df.index)+4, len(df.columns) - 1, {'type':'cell', 'criteria': '>', 'value':0, 'format':table_border_format})
#         merge_format = workbook.add_format({'align': 'center', 'bold': True, 'font_size': 16, 'border': 1})
#         worksheet.merge_range('A1:J2', f'National Water', merge_format)
#         merge_format = workbook.add_format({'align': 'center', 'bold': True, 'border': 1})
#         worksheet.merge_range('A3:J3', f'    Daily Collection Report   ', merge_format)
#         # worksheet.merge_range('E3:H3', f'Date: {def_date}', merge_format)
#         # worksheet.merge_range('I3:M3', f'Total bottle: {total_bottle}', merge_format)
#         merge_format = workbook.add_format({'align': 'center', 'bold': True, 'border': 1})
#         worksheet.merge_range('A4:J4', '', merge_format)
    
#     filename = f"Daily Collection Report.xlsx" 
#     response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'inline; filename = "{filename}"'
#     return response

#------------------Product-Route wise sales report


def product_route_salesreport(request):
    template = 'sales_management/product_route_salesreport.html'
    filter_data = {}

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    selected_product_id = request.GET.get('selected_product_id')

    selected_product = None
    if selected_product_id:
        selected_product = get_object_or_404(ProdutItemMaster, id=selected_product_id)
        filter_data["product_id"]=selected_product.pk

    query = request.GET.get("q")
    route_filter = request.GET.get('route_name')  # Modify to use POST instead of GET

    # Start with all customers
    user_li = Customers.objects.all()

    # Apply filters if they exist
    if query:
        user_li = user_li.filter(
            Q(custom_id__icontains=query) |
            Q(customer_name__icontains=query) |
            Q(sales_type__icontains=query) |

            Q(mobile_no__icontains=query) |
            Q(routes__route_name__icontains=query) |
            Q(location__location_name__icontains=query) |
            Q(building_name__icontains=query)
        )
        # print("user_li",user_li)

    if route_filter:
        user_li = user_li.filter(routes__route_name=route_filter)
        filter_data["route_name"]=route_filter

    # Get all route names for the dropdown
    route_li = RouteMaster.objects.all()

    customersupplyitems = CustomerSupplyItems.objects.all()
    coupons_collected = CustomerSupplyCoupon.objects.all()
    products = ProdutItemMaster.objects.all()
    today = datetime.today().date()

    if start_date and end_date:
        customersupplyitems = customersupplyitems.filter(customer_supply__created_date__range=[start_date, end_date], product=selected_product,customer_supply__customer__routes__route_name=route_filter)
        # coupons_collected = coupons_collected.filter(customer_supply__created_date__range=[start_date, end_date],customer_supply__customersales_type='CASH COUPON')
        coupons_collected = coupons_collected.filter(customer_supply__created_date__range=[start_date, end_date], customer_supply__customer__sales_type='CASH COUPON')
        filter_data['start_date'] = start_date
        filter_data['end_date'] = end_date
    else:
        customersupplyitems = customersupplyitems.filter(customer_supply__created_date=timezone.now().date(), product=selected_product,customer_supply__customer__routes__route_name=route_filter)
        coupons_collected = coupons_collected.filter(customer_supply__created_date=timezone.now().date(),customer_supply__customer__sales_type='CASH COUPON')
        filter_data['start_date'] = today
        filter_data['end_date'] = today

    context = {
        'customersupplyitems': customersupplyitems.order_by("-customer_supply__created_date"),
        'products': products,
        'today': today,
        'filter_data': filter_data,
        'coupons_collected': coupons_collected,
        'route_li': route_li,
    }
    return render(request, template, context)



def download_product_sales_excel(request):
    # Retrieve filter parameters from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    route_name = request.GET.get('route_name')
    selected_product_id = request.GET.get('selected_product_id')

    # Filter customer supply items based on the provided parameters
    customer_supply_items = CustomerSupplyItems.objects.filter(
        customer_supply__created_date__range=[start_date, end_date],
        product_id=selected_product_id,
        customer_supply__customer__routes__route_name=route_name
    ).order_by("-customer_supply__created_date")

    # Prepare data for Excel file
    data = {
        'Time of Supply': [],
        'Ref/Invoice No': [],
        'Route Name': [],
        'Customer Name': [],
        'Mode of Supply': [],
        'Quantity': [],
        'Empty Bottle Collected': [],
        'Coupon Collected': [],
        'Amount Collected': [],
    }

    for item in customer_supply_items:
        data['Time of Supply'].append(item.customer_supply.created_date.strftime('%d/%m/%Y'))
        data['Ref/Invoice No'].append(item.customer_supply.reference_number)
        data['Route Name'].append(item.customer_supply.customer.routes.route_name)
        data['Customer Name'].append(item.customer_supply.customer.customer_name)
        data['Mode of Supply'].append(item.customer_supply.customer.sales_type)
        data['Quantity'].append(item.quantity)
        data['Empty Bottle Collected'].append(item.customer_supply.collected_empty_bottle)
        data['Coupon Collected'].append(item.leaf_count())
        data['Amount Collected'].append(item.amount)

    # Create DataFrame
    df = pd.DataFrame(data)

    # Write DataFrame to Excel buffer
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Product Sales', index=False)

    # Prepare HTTP response with Excel file
    filename = f"Product_Sales_Report.xlsx"
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def download_product_sales_print(request):
    # Retrieve filter parameters from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    route_name = request.GET.get('route_name')
    selected_product_id = request.GET.get('selected_product_id')

    # Filter customer supply items based on the provided parameters
    customer_supply_items = CustomerSupplyItems.objects.filter(
        customer_supply__created_date__range=[start_date, end_date],
        product_id=selected_product_id,
        customer_supply__customer__routes__route_name=route_name
    ).order_by("-customer_supply__created_date")

    # Prepare data for PDF file
    data = [
        ["Time of Supply", "Ref/Invoice No", "Route Name", "Customer Name", "Mode of Supply", "Quantity",
         "Empty Bottle Collected", "Coupon Collected", "Amount Collected"]
    ]

    for item in customer_supply_items:
        data.append([
            item.customer_supply.created_date.strftime('%d/%m/%Y'),
            item.customer_supply.reference_number,
            item.customer_supply.customer.routes.route_name,
            item.customer_supply.customer.customer_name,
            item.customer_supply.customer.sales_type,
            item.quantity,
            item.customer_supply.collected_empty_bottle,
            item.leaf_count(),
            item.amount
        ])

 # Create PDF document with landscape orientation
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Product_Sales_Report.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(letter))  # Set orientation to landscape

    table = Table(data)

    # Add style to table
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Adjust bottom padding for header rows
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)

    # Add table to the PDF document
    doc.build([table])

    return response

# def yearmonthsalesreport(request):
#     user_li = Customers.objects.all()
#     # user_li = user_li.filter(routes__route_name=route_filter)
#     route_li = RouteMaster.objects.all()

#     context = {
#         'user_li': user_li, 
#         'route_li': route_li,
        
#             }
#     # route_li = RouteMaster.objects.all()
#     # print(route_li,'route_li')
#     # context = {'route_li': route_li}

#     return render(request,'sales_management/yearmonthsalesreport.html',context)
def yearmonthsalesreport(request):
    # Get all customers and routes
    user_li = Customers.objects.all()
    route_li = RouteMaster.objects.all()

    # Calculate YTD and MTD sales for each route
    route_sales = []
    current_year = datetime.now().year
    print(route_sales,"route_sales")
    print(current_year,'current_year')
    current_month = datetime.now().month
    print(current_month,'current_month')


    for route in route_li:
        ytd_sales = CustomerSupply.objects.filter(customer__routes=route, created_date__year=current_year).aggregate(total_sales=Sum('grand_total'))['total_sales'] or 0
        print('ytd_sales',ytd_sales)

        mtd_sales = CustomerSupply.objects.filter(customer__routes=route, created_date__year=current_year, created_date__month=current_month).aggregate(total_sales=Sum('grand_total'))['total_sales'] or 0
        print('mtd_sales',mtd_sales)

        route_sales.append({
            'route': route,
            'ytd_sales': ytd_sales,
            'mtd_sales': mtd_sales,
            'year': current_year,
        })
        print(route_sales,'route_sales')


    context = {
        'user_li': user_li, 
        'route_li' : route_li,
        'route_sales': route_sales,
    }

    return render(request, 'sales_management/yearmonthsalesreport.html', context)

# def yearmonthsalesreportview(request,route_id):
#     route = RouteMaster.objects.get(route_id = route_id)
#     print(route,'route')
#     route_customer = Customers.objects.filter(routes__route_name=route)
#     print(route_customer,'route_customer')

#     context = {
#         'route_customer': route_customer
#         }
    
#     return render(request, 'sales_management/yearmonthsalesreportview.html',context)
# def yearmonthsalesreportview(request, route_id):
#     route = RouteMaster.objects.get(route_id=route_id)
#     route_customers = Customers.objects.filter(routes__route_name=route)

#     # Calculate YTD and MTD sales for each customer
#     today = date.today()
#     ytd_sales = CustomerSupply.objects.filter(customer__in=route_customers, created_date__year=today.year).aggregate(total_ytd_sales=Sum('grand_total'))['total_ytd_sales'] or 0
#     mtd_sales = CustomerSupply.objects.filter(customer__in=route_customers, created_date__year=today.year, created_date__month=today.month).aggregate(total_mtd_sales=Sum('grand_total'))['total_mtd_sales'] or 0

#     # Associate YTD and MTD sales with each customer
#     customers_with_sales = []
#     for customer in route_customers:
#         ytd_customer_sales = CustomerSupply.objects.filter(customer=customer, created_date__year=today.year).aggregate(total_ytd_sales=Sum('grand_total'))['total_ytd_sales'] or 0
#         mtd_customer_sales = CustomerSupply.objects.filter(customer=customer, created_date__year=today.year, created_date__month=today.month).aggregate(total_mtd_sales=Sum('grand_total'))['total_mtd_sales'] or 0

#         customers_with_sales.append({
#             'customer': customer,
#             'ytd_sales': ytd_customer_sales,
#             'mtd_sales': mtd_customer_sales
#         })

#     context = {
#         'route': route,
#         'customers_with_sales': customers_with_sales,
#         'ytd_sales': ytd_sales,
#         'mtd_sales': mtd_sales
#     }

#     return render(request, 'sales_management/yearmonthsalesreportview.html', context)

def yearmonthsalesreportview(request, route_id):
    route = RouteMaster.objects.get(route_id=route_id)
    current_year = datetime.now().year

    yearly_sales = CustomerSupply.objects.filter(customer__routes=route, created_date__year=current_year).aggregate(total_sales=Sum('grand_total'))['total_sales'] or 0

    customers = Customers.objects.filter(routes__route_name=route)
    monthly_sales = []
    for customer in customers:
        monthly_sales_data = []
        for month in range(1, 13):
            month_date = datetime(current_year, month, 1)

            month_name = month_date.strftime("%B")  # Get month name
            monthly_sales_amount = CustomerSupply.objects.filter(customer=customer, created_date__year=current_year, created_date__month=month).aggregate(total_sales=Sum('grand_total'))['total_sales'] or 0
            monthly_sales_data.append({month_name: monthly_sales_amount})  # Append month name and sales amount
        monthly_sales.append({'customer': customer, 'monthly_sales': monthly_sales_data})

    context = {
        'route_id': route_id,
        'yearly_sales': yearly_sales,
        'monthly_sales': monthly_sales,
    }
    return render(request, 'sales_management/yearmonthsalesreportview.html', context)

#---------------------New Sales Report-----------------------------


def customerSales_report(request):
    filter_data = {}
    
    total_amount = 0
    total_discount = 0
    total_net_payable = 0
    total_vat = 0
    total_grand_total = 0
    total_amount_recieved = 0

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else :
        start_date = datetime.today().date()
        end_date = datetime.today().date()
    # print(start_date,end_date)
    filter_data = {
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
    }
    
    sales = CustomerSupply.objects.select_related('customer', 'salesman').filter(
        created_date__date__gte=start_date,  # Assuming start_date and end_date are provided
        created_date__date__lte=end_date
    ).exclude(customer__sales_type__in=["CASH COUPON","CREDIT"]).order_by("-created_date")

    # Query CustomerCoupon data
    coupons = CustomerCoupon.objects.select_related('customer', 'salesman').filter(
        created_date__date__gte=start_date,  # Assuming start_date and end_date are provided
        created_date__date__lte=end_date
    ).order_by("-created_date")

    # Query CollectionPayment data
    collections = CollectionPayment.objects.select_related('customer', 'salesman').filter(
        created_date__date__gte=start_date,  # Assuming start_date and end_date are provided
        created_date__date__lte=end_date
    ).order_by("-created_date")

    # Organize the data for rendering in the template
    sales_report_data = []

    # Process CustomerSupply data
    for sale in sales:
        sales_report_data.append({
            'date': sale.created_date.date(),
            'ref_invoice_no': sale.reference_number,
            'customer_name': sale.customer.customer_name,
            'building_name':sale.customer.building_name,
            'sales_type':sale.customer.sales_type,
            'route_name':sale.customer.routes.route_name,
            'salesman':sale.customer.sales_staff.get_fullname(),
            'amount': sale.grand_total,
            'discount': sale.discount,
            'net_taxable': sale.subtotal,
            'vat_amount': sale.vat,
            'grand_total': sale.grand_total,
            'amount_collected': sale.amount_recieved,
        })
        
        total_amount += sale.grand_total
        total_discount += sale.discount
        total_net_payable += sale.net_payable
        total_vat += sale.vat
        total_grand_total += sale.grand_total
        total_amount_recieved += sale.amount_recieved

    # Process CustomerCoupon data
    for coupon in coupons:
        sales_report_data.append({
            'date': coupon.created_date.date(),
            'ref_invoice_no': coupon.reference_number,
            'customer_name': coupon.customer.customer_name,
            'building_name':coupon.customer.building_name,
            'sales_type':coupon.customer.sales_type,
            'route_name':coupon.customer.routes.route_name,
            'salesman':coupon.customer.sales_staff.get_fullname(),
            # Add other fields as needed from CustomerCoupon model
            'amount': coupon.grand_total,
            'discount': coupon.discount,
            'net_taxable': coupon.net_amount,
            'vat_amount': Tax.objects.get(name="VAT").percentage,
            'grand_total': coupon.grand_total,
            'amount_collected': coupon.amount_recieved,
        })
        
        total_amount += coupon.grand_total
        total_discount += coupon.discount
        total_net_payable += coupon.net_amount
        total_vat += Tax.objects.get(name="VAT").percentage
        total_grand_total += coupon.grand_total
        total_amount_recieved += coupon.amount_recieved

    # Process CollectionPayment data
    for collection in collections:
        sales_report_data.append({
            'date': collection.created_date.date(),
            'ref_invoice_no': "",
            'customer_name': collection.customer.customer_name,
            'building_name':collection.customer.building_name,
            'sales_type':collection.customer.sales_type,
            'route_name':collection.customer.routes.route_name,
            'salesman':collection.customer.sales_staff.get_fullname(),
            # Add other fields as needed from CollectionPayment model
            'amount': collection.total_amount(),
            'discount': collection.total_discounts(),
            'net_taxable': collection.total_net_taxeble(),
            'vat_amount': collection.total_vat(),
            'grand_total': collection.total_amount(),
            'amount_collected': collection.collected_amount(),
            # Add other necessary data
        })
        
        total_amount += collection.total_amount()
        total_discount += collection.total_discounts()
        total_net_payable += collection.total_net_taxeble()
        total_vat += collection.total_vat()
        total_grand_total += collection.total_amount()
        total_amount_recieved += collection.collected_amount()

        
    context = {
        'customersales': sales_report_data,
        'total_amount': total_amount,
        'total_discount': total_discount,
        'total_net_payable': total_net_payable,
        'total_vat': total_vat,
        'total_grand_total': total_grand_total,
        'total_amount_recieved' :total_amount_recieved,

        'filter_data': filter_data,
        
    }
    return render(request, 'sales_management/customerSales_report.html', context)

def customerSales_Detail_report(request, id):
    customersale = get_object_or_404(CustomerSupplyItems, id=id)
    return render(request, 'sales_management/customerSales_Detail_report.html', {'customersale': customersale})



def customerSales_Excel_report(request):
    # Initialize totals
    total_amount = 0
    total_discount = 0
    total_net_payable = 0
    total_vat = 0
    total_grand_total = 0
    total_amount_received = 0

    # Get start and end dates from request parameters or default to today
    start_date = request.GET.get('start_date')
    # print("start_date", start_date)
    end_date = request.GET.get('end_date')
    # print("end_date", end_date)
    
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
        end_date = datetime.today().date()
    
    # Filter data dictionary
    filter_data = {
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
    }
    
    # Query CustomerSupply data
    sales = CustomerSupply.objects.select_related('customer', 'salesman').filter(
        created_date__date__gte=start_date,
        created_date__date__lte=end_date
    ).exclude(customer__sales_type__in=["CASH COUPON", "CREDIT"]).order_by("-created_date")

    # Query CustomerCoupon data
    coupons = CustomerCoupon.objects.select_related('customer', 'salesman').filter(
        created_date__date__gte=start_date,
        created_date__date__lte=end_date
    ).order_by("-created_date")

    # Query CollectionPayment data
    collections = CollectionPayment.objects.select_related('customer', 'salesman').filter(
        created_date__date__gte=start_date,
        created_date__date__lte=end_date
    ).order_by("-created_date")

    # Organize the data for rendering in the template
    sales_report_data = []

    # Process CustomerSupply data
    for idx, sale in enumerate(sales, start=1):
        sales_report_data.append({
            'sl_no': idx,
            'date': sale.created_date.date(),
            'ref_invoice_no': sale.reference_number,
            'customer_name': sale.customer.customer_name,
            'building_name': sale.customer.building_name,
            'sales_type': sale.customer.sales_type,
            'route_name': sale.customer.routes.route_name,
            'salesman': sale.customer.sales_staff.get_fullname(),
            'amount': sale.grand_total,
            'discount': sale.discount,
            'net_taxable': sale.subtotal,
            'vat_amount': sale.vat,
            'grand_total': sale.grand_total,
            'amount_collected': sale.amount_recieved,
        })
        
        total_amount += sale.grand_total
        total_discount += sale.discount
        total_net_payable += sale.net_payable
        total_vat += sale.vat
        total_grand_total += sale.grand_total
        total_amount_received += sale.amount_recieved

    # Process CustomerCoupon data
    for idx, coupon in enumerate(coupons, start=len(sales_report_data) + 1):
        sales_report_data.append({
            'sl_no': idx,
            'date': coupon.created_date.date(),
            'ref_invoice_no': coupon.reference_number,
            'customer_name': coupon.customer.customer_name,
            'building_name': coupon.customer.building_name,
            'sales_type': coupon.customer.sales_type,
            'route_name': coupon.customer.routes.route_name,
            'salesman': coupon.customer.sales_staff.get_fullname(),
            'amount': coupon.grand_total,
            'discount': coupon.discount,
            'net_taxable': coupon.net_amount,
            'vat_amount': Tax.objects.get(name="VAT").percentage,
            'grand_total': coupon.grand_total,
            'amount_collected': coupon.amount_recieved,
        })
        
        total_amount += coupon.grand_total
        total_discount += coupon.discount
        total_net_payable += coupon.net_amount
        total_vat += Tax.objects.get(name="VAT").percentage
        total_grand_total += coupon.grand_total
        total_amount_received += coupon.amount_recieved

    # Process CollectionPayment data
    for idx, collection in enumerate(collections, start=len(sales_report_data) + 1):
        sales_report_data.append({
            'sl_no': idx,
            'date': collection.created_date.date(),
            'ref_invoice_no': "",
            'customer_name': collection.customer.customer_name,
            'building_name': collection.customer.building_name,
            'sales_type': collection.customer.sales_type,
            'route_name': collection.customer.routes.route_name,
            'salesman': collection.customer.sales_staff.get_fullname(),
            'amount': collection.total_amount(),
            'discount': collection.total_discounts(),
            'net_taxable': collection.total_net_taxeble(),
            'vat_amount': collection.total_vat(),
            'grand_total': collection.total_amount(),
            'amount_collected': collection.collected_amount(),
        })
        
        total_amount += collection.total_amount()
        total_discount += collection.total_discounts()
        total_net_payable += collection.total_net_taxeble()
        total_vat += collection.total_vat()
        total_grand_total += collection.total_amount()
        total_amount_received += collection.collected_amount()

    # Create the HttpResponse object with Excel content type and attachment filename
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="customerSales_Excel_report.xlsx"'

    # Create a new Excel workbook and add a worksheet
    workbook = xlsxwriter.Workbook(response)
    worksheet = workbook.add_worksheet()

    # Define cell formats
    bold_format = workbook.add_format({'bold': True})

    # Write headers
    headers = ['Sl No', 'Ref/Invoice No', 'Date & Time', 'Customer Name', 'Building Name/Room No/Floor No','SalesType', 'Route', 'Salesman', 'Amount', 'Discount', 'Net Taxable', 'Vat Amount', 'Grand Total','Amount Collected']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, bold_format)

    # Write data rows
    for row, sale_data in enumerate(sales_report_data, start=1):
        worksheet.write(row, 0, sale_data['sl_no'])
        worksheet.write(row, 1, sale_data['ref_invoice_no'])
        worksheet.write(row, 2, str(sale_data['date']))
        worksheet.write(row, 3, sale_data['customer_name'])
        worksheet.write(row, 4, sale_data['building_name'])
        worksheet.write(row, 5, sale_data['sales_type'])
        worksheet.write(row, 6, sale_data['route_name'])
        worksheet.write(row, 7, sale_data['salesman'])
        worksheet.write(row, 8, sale_data['amount'])
        worksheet.write(row, 9, sale_data['discount'])
        worksheet.write(row, 10, sale_data['net_taxable'])
        worksheet.write(row, 11, sale_data['vat_amount'])
        worksheet.write(row, 12, sale_data['grand_total'])
        worksheet.write(row, 13, sale_data['amount_collected'])

    # Write totals
    total_row = len(sales_report_data) + 1
    worksheet.write(total_row, 7, 'Total:')
    worksheet.write(total_row, 8, total_amount)
    worksheet.write(total_row, 9, total_discount)
    worksheet.write(total_row, 10, total_net_payable)
    worksheet.write(total_row, 11, total_vat)
    worksheet.write(total_row, 12, total_grand_total)
    worksheet.write(total_row, 13, total_amount_received)

    # Close the workbook
    workbook.close()

    return response



def customerSales_Print_report(request):
    # Initialize totals
    total_amount = 0
    total_discount = 0
    total_net_payable = 0
    total_vat = 0
    total_grand_total = 0
    total_amount_received = 0

    # Get start and end dates from request parameters or default to today
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
        end_date = datetime.today().date()
    
    # Query CustomerSupply data
    sales = CustomerSupply.objects.select_related('customer', 'salesman').filter(
        created_date__date__gte=start_date,
        created_date__date__lte=end_date
    ).exclude(customer__sales_type__in=["CASH COUPON", "CREDIT"]).order_by("-created_date")

    # Query CustomerCoupon data
    coupons = CustomerCoupon.objects.select_related('customer', 'salesman').filter(
        created_date__date__gte=start_date,
        created_date__date__lte=end_date
    ).order_by("-created_date")

    # Query CollectionPayment data
    collections = CollectionPayment.objects.select_related('customer', 'salesman').filter(
        created_date__date__gte=start_date,
        created_date__date__lte=end_date
    ).order_by("-created_date")

    # Organize the data for rendering in the template
    sales_report_data = []

    # Process CustomerSupply data
    for idx, sale in enumerate(sales, start=1):
        sales_report_data.append({
            'date': sale.created_date.date(),
            'ref_invoice_no': sale.reference_number,
            'customer_name': sale.customer.customer_name,
            'building_name': sale.customer.building_name,
            'sales_type': sale.customer.sales_type,
            'route_name': sale.customer.routes.route_name,
            'salesman': sale.customer.sales_staff.get_fullname(),
            'amount': sale.grand_total,
            'discount': sale.discount,
            'net_taxable': sale.subtotal,
            'vat_amount': sale.vat,
            'grand_total': sale.grand_total,
            'amount_collected': sale.amount_recieved,
        })
        
        total_amount += sale.grand_total
        total_discount += sale.discount
        total_net_payable += sale.net_payable
        total_vat += sale.vat
        total_grand_total += sale.grand_total
        total_amount_received += sale.amount_recieved

    # Process CustomerCoupon data
    for idx, coupon in enumerate(coupons, start=len(sales_report_data) + 1):
        sales_report_data.append({
            'date': coupon.created_date.date(),
            'ref_invoice_no': coupon.reference_number,
            'customer_name': coupon.customer.customer_name,
            'building_name': coupon.customer.building_name,
            'sales_type': coupon.customer.sales_type,
            'route_name': coupon.customer.routes.route_name,
            'salesman': coupon.customer.sales_staff.get_fullname(),
            'amount': coupon.grand_total,
            'discount': coupon.discount,
            'net_taxable': coupon.net_amount,
            'vat_amount': Tax.objects.get(name="VAT").percentage,
            'grand_total': coupon.grand_total,
            'amount_collected': coupon.amount_recieved,
        })
        
        total_amount += coupon.grand_total
        total_discount += coupon.discount
        total_net_payable += coupon.net_amount
        total_vat += Tax.objects.get(name="VAT").percentage
        total_grand_total += coupon.grand_total
        total_amount_received += coupon.amount_recieved

    # Process CollectionPayment data
    for idx, collection in enumerate(collections, start=len(sales_report_data) + 1):
        sales_report_data.append({
            'date': collection.created_date.date(),
            'ref_invoice_no': "",
            'customer_name': collection.customer.customer_name,
            'building_name': collection.customer.building_name,
            'sales_type': collection.customer.sales_type,
            'route_name': collection.customer.routes.route_name,
            'salesman': collection.customer.sales_staff.get_fullname(),
            'amount': collection.total_amount(),
            'discount': collection.total_discounts(),
            'net_taxable': collection.total_net_taxeble(),
            'vat_amount': collection.total_vat(),
            'grand_total': collection.total_amount(),
            'amount_collected': collection.collected_amount(),
        })
        
        total_amount += collection.total_amount()
        total_discount += collection.total_discounts()
        total_net_payable += collection.total_net_taxeble()
        total_vat += collection.total_vat()
        total_grand_total += collection.total_amount()
        total_amount_received += collection.collected_amount()

    # Create a PDF report
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="customerSales_PDF_report.pdf"'

    # Create a PDF document
    # doc = SimpleDocTemplate(response, pagesize=landscape(A4))

    # Define custom page size (e.g., 8.5x5.5 inches)
    custom_page_size = (14.5 * inch, 5.5 * inch)

    # Create a PDF document with custom page size
    doc = SimpleDocTemplate(response, pagesize=landscape(custom_page_size))
    elements = []

    # Define styles
    styles = {
        'header': {'fontSize': 10, 'bold': True},
        'cell': {'fontSize': 8},
    }

    # Create table data
    data = []
    headers = ['Date & Time', 'Ref/Invoice No', 'Customer Name', 'Building Name', 'SalesType', 'Route', 'Salesman', 'Amount', 'Discount', 'Net Taxable', 'Vat Amount', 'Grand Total', 'Amount Collected']
    data.append(headers)
    for sale_data in sales_report_data:
        row = [sale_data['date'], sale_data['ref_invoice_no'], sale_data['customer_name'], sale_data['building_name'], sale_data['sales_type'], sale_data['route_name'], sale_data['salesman'], sale_data['amount'], sale_data['discount'], sale_data['net_taxable'], sale_data['vat_amount'], sale_data['grand_total'], sale_data['amount_collected']]
        data.append(row)

    # Create table and apply styles
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
    ]))
    elements.append(table)

  # Write totals
    totals_table = Table([
        ['Total:', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 
         '', '', '', '', '', '', '', '','', '', '', '', '', '', '', '', '', '',total_amount,'',total_discount, '', '', total_net_payable, '', '',total_vat, '', '', total_grand_total, '', '', '', total_amount_received]
    ])
    totals_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(totals_table)

    # Add elements to the document
    doc.build(elements)

    return response

#------------------Collection Report-------------------------                

# def collectionreport(request):
#     # Retrieve collection payments along with related fields
#     collection_payments = CollectionItems.objects.select_related('invoice','collection_payment__customer', 'collection_payment__customer__routes').all()

#     context = {
#         'collection_payments': collection_payments
#     }

#     return render(request, 'sales_management/collection_report.html', context)
def collectionreport(request):
    start_date = None
    end_date = None
    selected_date = None
    selected_route_id = None
    selected_route = None
    template = 'sales_management/collection_report.html'
    
    collection_payments = CollectionItems.objects.all()
    
    routes = RouteMaster.objects.all()
    route_counts = {}
    today = datetime.today()
    
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
        end_date = datetime.today().date() + timedelta(days=1)

    filter_data = {
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
    }
            
    collection_payments = collection_payments.filter(collection_payment__created_date__range=[start_date, end_date])
    
    if selected_route_id:
        selected_route = RouteMaster.objects.get(route_name=selected_route_id)
        collection_payments = collection_payments.filter(collection_payment__customer__routes__route_name=selected_route)
        filter_data = {'selected_route': selected_route}
    
    context = {
        'collection_payments': collection_payments, 
        'routes': routes, 
        'route_counts': route_counts, 
        'today': today,
        'filter_data': filter_data,
    }
    return render(request, template, context)

def collection_report_excel(request):
    instances = CollectionItems.objects.all()
    data = {
        'Date': [instance.collection_payment.created_date.date() for instance in instances],
        'Customer name': [instance.collection_payment.customer.customer_name for instance in instances],
        'Mobile No': [instance.collection_payment.customer.mobile_no for instance in instances],
        'Route': [instance.collection_payment.customer.routes.route_name if instance.collection_payment.customer.routes else '' for instance in instances],
        'Building Name': [instance.collection_payment.customer.building_name for instance in instances],
        'House No': [instance.collection_payment.customer.door_house_no if instance.collection_payment.customer.door_house_no else 'Nil' for instance in instances],
        'Receipt No/Reference No': [instance.invoice.reference_no for instance in instances],
        'Amount': [instance.amount for instance in instances],
        'Mode of Payment': [instance.collection_payment.payment_method for instance in instances],
    }
    df = pd.DataFrame(data)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=4)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        table_border_format = workbook.add_format({'border':1})
        worksheet.conditional_format(4, 0, len(df.index)+4, len(df.columns) - 1, {'type':'cell', 'criteria': '>', 'value':0, 'format':table_border_format})
        merge_format = workbook.add_format({'align': 'center', 'bold': True, 'font_size': 16, 'border': 1})
        worksheet.merge_range('A1:J2', f'National Water', merge_format)
        merge_format = workbook.add_format({'align': 'center', 'bold': True, 'border': 1})
        worksheet.merge_range('A3:J3', f'    Collection Report   ', merge_format)
        merge_format = workbook.add_format({'align': 'center', 'bold': True, 'border': 1})
        worksheet.merge_range('A4:J4', '', merge_format)
    
    filename = f"Collection Report.xlsx"
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'inline; filename = "{filename}"'
    return response




#-----------------Suspense Report--------------------------
from .forms import SuspenseCollectionForm

def suspense_report(request):
    start_date = request.GET.get('start_date')
    print("start_date", start_date)
    filter_data = {}

    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()

    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')

    invoices = InvoiceDailyCollection.objects.filter(created_date__gte=start_date)\
                                              .select_related('customer__routes')\
                                              .select_related('customer__sales_staff')
    print("invoices", invoices)
    suspense_collections = SuspenseCollection.objects.filter(created_date__gte=start_date)
    print("suspense_collections", suspense_collections)

    # Calculate total opening suspense
    total_opening_suspense = invoices.aggregate(total_opening_suspense=Sum('amount'))['total_opening_suspense'] or 0

    # Calculate total paid
    total_paid = suspense_collections.aggregate(total_paid=Sum('amount_paid'))['total_paid'] or 0

    # Calculate total closing suspense
    total_closing_suspense = suspense_collections.aggregate(total_closing_suspense=Sum('amount_balance'))['total_closing_suspense'] or 0

    context = {
        'invoices': invoices,
        'filter_data': filter_data,
        'suspense_collections': suspense_collections,
        'total_opening_suspense': total_opening_suspense,
        'total_paid': total_paid,
        'total_closing_suspense': total_closing_suspense,
    }

    return render(request, 'sales_management/suspense_report.html', context)

def create_suspense_collection(request, id):
    invoice = get_object_or_404(InvoiceDailyCollection, id=id)
    
    if request.method == 'POST':
        form = SuspenseCollectionForm(request.POST)
        if form.is_valid():
            # Fetch the total expense for the given route and van (assuming they are available in the invoice)
            total_expense = Expense.objects.filter(
                route=invoice.customer.routes,
                expense_date=invoice.created_date.date()  # Assuming expense_date is a DateField
            ).aggregate(total_amount=Sum('amount'))['total_amount'] or 0
            
            # Calculate net_payeble_amount
            cash_sale_amount = invoice.invoice.amout_total if invoice.invoice.invoice_type == 'cash_invoice' else 0
            credit_sale_amount = invoice.invoice.amout_total if invoice.invoice.invoice_type == 'credit_invoice' else 0
            net_payeble_amount = cash_sale_amount + credit_sale_amount - total_expense

            suspense_collection = form.save(commit=False)
            suspense_collection.date = timezone.now()  # Set the created_date
            suspense_collection.created_date = invoice.created_date 
            suspense_collection.salesman = invoice.customer.sales_staff
            suspense_collection.route = invoice.customer.routes
            suspense_collection.cash_sale_amount = cash_sale_amount
            suspense_collection.credit_sale_amount = credit_sale_amount
            suspense_collection.expense = total_expense
            suspense_collection.net_payeble_amount = net_payeble_amount  # Set the net_payeble_amount field
            # Calculate amount_balance
            amount_paid = form.cleaned_data['amount_paid']
            amount_balance = invoice.amount - amount_paid
            suspense_collection.amount_balance = amount_balance
            suspense_collection.save()
            print("SuspenseCollection instance saved successfully:", suspense_collection)

            return redirect('suspense_report')  # Redirect to the suspense_report URL
        else:
            print("Form errors:", form.errors)

    else:
        form = SuspenseCollectionForm()
        print("form",form)
    
    return render(request, 'sales_management/create_suspense_collection.html', {'form': form, 'invoice': invoice})


from django.utils import timezone

def suspense_report_excel(request):
    start_date = request.GET.get('start_date')
    filter_data = {}

    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()

    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')

    invoices = InvoiceDailyCollection.objects.filter(created_date__gte=start_date)\
                                              .select_related('customer__routes')\
                                              .select_related('customer__sales_staff')

    suspense_collections = SuspenseCollection.objects.filter(created_date__gte=start_date)

    # Create a BytesIO object to save workbook data
    buffer = BytesIO()

    # Create a new Excel workbook and add a worksheet
    workbook = xlsxwriter.Workbook(buffer)
    worksheet = workbook.add_worksheet()
    worksheet.title = "Suspense Report"

    # Write header row
    headers = ['Sl No', 'Created Date', 'Invoice Type', 'Route', 'Salesman', 'Opening Suspense', 'Paid', 'Closing Suspense']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)

    # Write data rows
    row = 1
    for invoice in invoices:
        created_date = invoice.created_date.replace(tzinfo=None)
        worksheet.write(row, 0, row)  # Sl No
        worksheet.write(row, 1, created_date)  # Created Date
        worksheet.write(row, 2, invoice.invoice.invoice_type)  # Invoice Type
        worksheet.write(row, 4, invoice.customer.sales_staff.username)  # Salesman
        worksheet.write(row, 5, invoice.amount)  # Opening Suspense
        # Fetch the corresponding SuspenseCollection for this invoice
        related_suspense_collection = suspense_collections.filter(salesman=invoice.customer.sales_staff).first()
        if related_suspense_collection:
            worksheet.write(row, 6, related_suspense_collection.amount_paid)  # Paid
            worksheet.write(row, 7, related_suspense_collection.amount_balance)  # Closing Suspense
        else:
            worksheet.write(row, 6, '')  # Paid
            worksheet.write(row, 7, '')  # Closing Suspense

        # Extract and write route information
        if invoice.customer.routes:
            route_info = f"{invoice.customer.routes.route_name}"
            worksheet.write(row, 3, route_info)  # Route
        else:
            worksheet.write(row, 3, '')  # Route

        row += 1

    # Close the workbook
    workbook.close()

    # Create HttpResponse object to return the Excel file as a response
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Suspense_Report.xlsx'
    return response


def suspense_report_print(request):
    start_date = request.GET.get('start_date')
    filter_data = {}

    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()

    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')

    invoices = InvoiceDailyCollection.objects.filter(created_date__gte=start_date)\
                                              .select_related('customer__routes')\
                                              .select_related('customer__sales_staff')

    suspense_collections = SuspenseCollection.objects.filter(created_date__gte=start_date)

    # Prepare data for the PDF report
    data = [['Sl No', 'Created Date', 'Invoice Type', 'Route', 'Salesman', 'Opening Suspense', 'Paid', 'Closing Suspense']]

    for invoice in invoices:
        created_date = invoice.created_date.replace(tzinfo=None)
        salesman_name = invoice.customer.sales_staff.username if invoice.customer.sales_staff else ''
        route_info = f"{invoice.customer.routes.route_name}" if invoice.customer.routes else ''

        related_suspense_collection = suspense_collections.filter(salesman=invoice.customer.sales_staff).first()
        paid = related_suspense_collection.amount_paid if related_suspense_collection else ''
        balance = related_suspense_collection.amount_balance if related_suspense_collection else ''

        data.append([invoice.id, created_date, invoice.invoice.invoice_type, route_info, salesman_name, invoice.amount, paid, balance])

    # Create a BytesIO object to save PDF data
    buffer = BytesIO()

    # Create a new PDF document
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    
    # Create a table with the data
    table = Table(data)
    
    # Add style to the table
    style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), 'grey'),
                        ('TEXTCOLOR', (0, 0), (-1, 0), 'white'),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), 'lightgrey'),
                        ('GRID', (0, 0), (-1, -1), 1, 'black')])

    table.setStyle(style)
    
    # Add table to the document
    doc.build([table])

    # Create HttpResponse object to return the PDF file as a response
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=Suspense_Report.pdf'
    return response


#-----------------DSR cash sales Report--------------------------

def cashsales_report(request):
    start_date = request.GET.get('start_date')
    filter_data = {}
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
    
    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')
    
    invoices = Invoice.objects.filter(created_date__date=start_date, invoice_type="cash_invoice")
    
    total_net_taxable = invoices.aggregate(total_net_taxable=Sum('net_taxable'))['total_net_taxable'] or 0
    total_vat = invoices.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
    total_amout_total = invoices.aggregate(total_amout_total=Sum('amout_total'))['total_amout_total'] or 0

    context = {
        'invoices': invoices,
        'filter_data': filter_data,
        
        'total_net_taxable':total_net_taxable,
        'total_vat':total_vat,
        'total_amout_total':total_amout_total,

    }
    return render(request, 'sales_management/dsr_cash_sales_report.html', context)

def cashsales_report_excel(request):
    start_date = request.GET.get('start_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
    
    invoices = Invoice.objects.filter(created_date__date=start_date, invoice_type="cash_invoice")

    # Calculate total values for Opening Suspense, Paid, and Closing Suspense
    total_net_taxable = sum(invoice.net_taxable for invoice in invoices)
    total_vat = sum(invoice.vat for invoice in invoices)
    total_amout_total = sum(invoice.amout_total for invoice in invoices)

    # Create a BytesIO object to save workbook data
    buffer = BytesIO()

    # Create a new Excel workbook and add a worksheet
    workbook = xlsxwriter.Workbook(buffer)
    worksheet = workbook.add_worksheet()
    worksheet.title = "Cash Sales Report"

    # Write headers
    headers = ["Sl No", "Reference No", "Customer Name", "Building Name", "Net taxable", "Vat","Grand Total"]
    worksheet.write_row(0, 0, headers)

    # Write data rows
    for index, invoice in enumerate(invoices, start=1):
        worksheet.write_row(index, 0, [
            index,
            invoice.reference_no,
            invoice.customer.customer_name,
            invoice.customer.building_name,
            invoice.net_taxable,
            invoice.vat,
            invoice.amout_total
        ])

    # Write footer data
    footer_row = len(invoices) + 1
    worksheet.write(footer_row, 3, "Total")
    worksheet.write(footer_row, 4, total_net_taxable)
    worksheet.write(footer_row, 5, total_vat)
    worksheet.write(footer_row, 6, total_amout_total)

    # Close the workbook
    workbook.close()

    # Create HttpResponse object to return the Excel file as a response
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Cash_Sales_Report.xlsx'
    return response


def cashsales_report_print(request):
    start_date = request.GET.get('start_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
    
    invoices = Invoice.objects.filter(created_date__date=start_date, invoice_type="cash_invoice")

    total_net_taxable = sum(invoice.net_taxable for invoice in invoices)
    total_vat = sum(invoice.vat for invoice in invoices)
    total_amout_total = sum(invoice.amout_total for invoice in invoices)

    # Create a BytesIO object to save PDF data
    buffer = BytesIO()

    # Create a new PDF document
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    
    # Define table data
    data = [
        ["Sl No", "Reference No", "Customer Name", "Building Name", "Net taxable", "Vat","Grand Total"]
    ]
    
    # Populate table data
    for index, invoice in enumerate(invoices, start=1):
        data.append([
            str(index),
            str(invoice.reference_no),
            str(invoice.customer.customer_name),
            str(invoice.customer.building_name),
            str(invoice.net_taxable),
            str(invoice.vat),
            str(invoice.amout_total)
        ])
    
    # Add footer data
    data.append([
        "", "","", "Total", str(total_net_taxable), str(total_vat), str(total_amout_total)
    ])

    # Create a table
    table = Table(data)

    # Add style to the table
    style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ])
    table.setStyle(style)

    # Add table to the document
    doc.build([table])

    # Create HttpResponse object to return the PDF file as a response
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=Cash_Sales_Report.pdf'
    return response



#-----------------DSR credit sales Report--------------------------

def creditsales_report(request):
    start_date = request.GET.get('start_date')
    filter_data = {}
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
    
    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')
    
    invoices = Invoice.objects.filter(created_date__date=start_date, invoice_type="credit_invoive")
    
    total_net_taxable = invoices.aggregate(total_net_taxable=Sum('net_taxable'))['total_net_taxable'] or 0
    total_vat = invoices.aggregate(total_vat=Sum('vat'))['total_vat'] or 0
    total_amout_total = invoices.aggregate(total_amout_total=Sum('amout_total'))['total_amout_total'] or 0

    context = {
        'invoices': invoices,
        'filter_data': filter_data,
        
        'total_net_taxable':total_net_taxable,
        'total_vat':total_vat,
        'total_amout_total':total_amout_total,

    }
    return render(request, 'sales_management/dsr_credit_sales_report.html', context)

def creditsales_report_excel(request):
    start_date = request.GET.get('start_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
    
    invoices = Invoice.objects.filter(created_date__date=start_date, invoice_type="credit_invoice")

    # Calculate total values for Opening Suspense, Paid, and Closing Suspense
    total_net_taxable = sum(invoice.net_taxable for invoice in invoices)
    total_vat = sum(invoice.vat for invoice in invoices)
    total_amout_total = sum(invoice.amout_total for invoice in invoices)

    # Create a BytesIO object to save workbook data
    buffer = BytesIO()

    # Create a new Excel workbook and add a worksheet
    workbook = xlsxwriter.Workbook(buffer)
    worksheet = workbook.add_worksheet()
    worksheet.title = "Credit Sales Report"

    # Write headers
    headers = ["Sl No", "Reference No", "Customer Name", "Building Name", "Net taxable", "Vat","Grand Total"]
    worksheet.write_row(0, 0, headers)

    # Write data rows
    for index, invoice in enumerate(invoices, start=1):
        worksheet.write_row(index, 0, [
            index,
            invoice.reference_no,
            invoice.customer.customer_name,
            invoice.customer.building_name,
            invoice.net_taxable,
            invoice.vat,
            invoice.amout_total
        ])

    # Write footer data
    footer_row = len(invoices) + 1
    worksheet.write(footer_row, 3, "Total")
    worksheet.write(footer_row, 4, total_net_taxable)
    worksheet.write(footer_row, 5, total_vat)
    worksheet.write(footer_row, 6, total_amout_total)

    # Close the workbook
    workbook.close()

    # Create HttpResponse object to return the Excel file as a response
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Credit_Sales_Report.xlsx'
    return response


def creditsales_report_print(request):
    start_date = request.GET.get('start_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
    
    invoices = Invoice.objects.filter(created_date__date=start_date, invoice_type="credit_invoice")

    total_net_taxable = sum(invoice.net_taxable for invoice in invoices)
    total_vat = sum(invoice.vat for invoice in invoices)
    total_amout_total = sum(invoice.amout_total for invoice in invoices)

    # Create a BytesIO object to save PDF data
    buffer = BytesIO()

    # Create a new PDF document
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    
    # Define table data
    data = [
        ["Sl No", "Reference No", "Customer Name", "Building Name", "Net taxable", "Vat","Grand Total"]
    ]
    
    # Populate table data
    for index, invoice in enumerate(invoices, start=1):
        data.append([
            str(index),
            str(invoice.reference_no),
            str(invoice.customer.customer_name),
            str(invoice.customer.building_name),
            str(invoice.net_taxable),
            str(invoice.vat),
            str(invoice.amout_total)
        ])
    
    # Add footer data
    data.append([
        "", "","", "Total", str(total_net_taxable), str(total_vat), str(total_amout_total)
    ])

    # Create a table
    table = Table(data)

    # Add style to the table
    style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ])
    table.setStyle(style)

    # Add table to the document
    doc.build([table])

    # Create HttpResponse object to return the PDF file as a response
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=Credit_Sales_Report.pdf'
    return response

#----------DSR Coupon Book Sales--------------

def dsr_coupon_book_sales(request):

    # Get start date from request parameters or default to today
    start_date = request.GET.get('start_date')
    print("start_date",start_date)
    
    filter_data = {}
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
    
    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')
    customer_coupon_items=CustomerCouponItems.objects.filter(customer_coupon__created_date__gte=start_date).order_by("-customer_coupon__created_date")

    context={'customer_coupon_items':customer_coupon_items,
            'filter_data': filter_data,
}
    return render(request,'sales_management/dsr_coupon_book_sales.html',context)

def dsr_coupon_book_sales_excel(request):
    # Get start date from request parameters or default to today
    start_date = request.GET.get('start_date')
    print("start_date", start_date)
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
    
    # Query data
    customer_coupon_items = CustomerCouponItems.objects.filter(customer_coupon__created_date__gte=start_date).order_by("-customer_coupon__created_date")
    
    # Create a BytesIO object to save workbook data
    buffer = BytesIO()

    # Create a new Excel workbook and add a worksheet
    workbook = xlsxwriter.Workbook(buffer)
    worksheet = workbook.add_worksheet()

    # Write headers
    headers = ["Sl no","Reference Number", "Customer Name", "Building Name/no", "Book Number", "Coupon Type" , "Coupon Method","Rate"]
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)

    # Write data to the worksheet
    row = 1  # Start from row 1 after the header row
    sl_no = 1  # Initialize Sl No
    for item in customer_coupon_items:
        building_info = f"{item.customer_coupon.customer.building_name} / {item.customer_coupon.customer.door_house_no} / {item.customer_coupon.customer.floor_no}"
        worksheet.write(row, 0, sl_no)  
        worksheet.write(row, 1, item.customer_coupon.reference_number)
        worksheet.write(row, 2, item.customer_coupon.customer.customer_name)
        worksheet.write(row, 3, building_info)
        worksheet.write(row, 4, item.coupon.book_num)
        worksheet.write(row, 5, item.coupon.coupon_type.coupon_type_name)
        worksheet.write(row, 6, item.coupon.coupon_method)
        worksheet.write(row, 7, item.rate)
        row += 1
        sl_no += 1  # Increment Sl No

    # Close the workbook
    workbook.close()

    # Move the BytesIO pointer to the start of the buffer
    buffer.seek(0)

    # Create the HTTP response with the Excel file
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="dsr_coupon_book_sales.xlsx"'

    return response

def dsr_coupon_book_sales_print(request):
    # Get start date from request parameters or default to today
    start_date = request.GET.get('start_date')
    print("start_date", start_date)
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
    
    # Query data
    customer_coupon_items = CustomerCouponItems.objects.filter(customer_coupon__created_date__gte=start_date).order_by("-customer_coupon__created_date")

    # Create data for PDF table
    data = [
        ["Reference Number", "Customer Name", "Building Name/no", "Book Number", "Coupon Type", "Coupon Method","Rate"]
    ]
    for item in customer_coupon_items:
        building_info = f"{item.customer_coupon.customer.building_name} / {item.customer_coupon.customer.door_house_no} / {item.customer_coupon.customer.floor_no}"
        data.append([
            item.customer_coupon.reference_number,
            item.customer_coupon.customer.customer_name,
            building_info,
            item.coupon.book_num,
            item.coupon.coupon_type.coupon_type_name,
            item.coupon.coupon_method,
            item.rate
        ])

    # Define custom page size (e.g., 8.5x5.5 inches)
    custom_page_size = (13.5 * inch, 5.5 * inch)

    # Create a PDF document with custom page size
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(custom_page_size))
    table = Table(data, colWidths=[100, 150, 200, 100, 150, 100])
    table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                               ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                               ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                               ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                               ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                               ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                               ('GRID', (0, 0), (-1, -1), 1, colors.black)]))
    doc.build([table])

    # Create HTTP response with the PDF file
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="dsr_coupon_book_sales.pdf"'
    return response 


#-----------------DSR Stock Report--------------------------

def dsr_stock_report(request):
    start_date = request.GET.get('start_date')
    filter_data = {}
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
    
    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')
    salesman_id=request.user.id
    products = ProdutItemMaster.objects.filter()
    van_instances = Van.objects.filter(salesman=salesman_id,created_date__date=start_date)
    van_product_stock = VanProductStock.objects.filter(van__created_date__date=start_date)
    

    context = {
        'products': products,
        'van_instances': van_instances,
        'van_product_stock': van_product_stock,
        

    }
    return render(request, 'sales_management/dsr_stock_report.html', context)

def visitstatistics_report(request):
    user_id = request.user.id
    start_date = request.GET.get('start_date')
    filter_data = {}
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()

    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')
    #new customers created
    salesman_customers_count = Customers.objects.filter(created_date__date=start_date,sales_staff_id=user_id).count()
    #emergency supply
    emergency_customers = DiffBottlesModel.objects.filter(created_date__date=start_date, assign_this_to_id=user_id).count()
    print('emergency_customers',emergency_customers)
    #actual visit
    visited_customers_count = CustomerSupply.objects.filter(salesman_id=user_id, created_date__date=start_date).count()
    print(visited_customers_count,'visited_customers_count')
    
    planned_visits_count = Customers.objects.filter(visit_schedule__isnull=False, created_date__date=start_date).annotate(
        planned_visit_count=Count('visit_schedule')
    ).filter(planned_visit_count__gt=0).count()
    print(planned_visits_count,'planned_visits_count')

    # non_visited_customers = CustomerSupply.objects.exclude(salesman_id=user_id, created_date__date=start_date).count()
    non_visited_customers = planned_visits_count - visited_customers_count

    print(non_visited_customers,'non_visited_customers')

    context = {
        'salesman_customers_count': salesman_customers_count,
        'emergency_customers': emergency_customers,
        'visited_customers_count': visited_customers_count,
        'non_visited_customers': non_visited_customers,
        'planned_visits_count': planned_visits_count
    }
    return render(request, 'sales_management/dsr_visit_statistics_report.html', context)


def visitstatistics_report_excel(request):
    user_id = request.user.id

    start_date = request.GET.get('start_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
    
    salesman_customers_count = Customers.objects.filter(created_date__date=start_date,sales_staff_id=user_id).count()
    #emergency supply
    emergency_customers = DiffBottlesModel.objects.filter(created_date__date=start_date, assign_this_to_id=user_id).count()
    print('emergency_customers',emergency_customers)
    #actual visit
    visited_customers_count = CustomerSupply.objects.filter(salesman_id=user_id, created_date__date=start_date).count()
    print(visited_customers_count,'visited_customers_count')
    
    planned_visits_count = Customers.objects.filter(visit_schedule__isnull=False, created_date__date=start_date).annotate(
        planned_visit_count=Count('visit_schedule')
    ).filter(planned_visit_count__gt=0).count()
    print(planned_visits_count,'planned_visits_count')

    # non_visited_customers = CustomerSupply.objects.exclude(salesman_id=user_id, created_date__date=start_date).count()
    non_visited_customers = planned_visits_count - visited_customers_count

   

    # Create a BytesIO object to save workbook data
    buffer = BytesIO()

    # Create a new Excel workbook and add a worksheet
    workbook = xlsxwriter.Workbook(buffer)
    worksheet = workbook.add_worksheet()
    worksheet.title = "Credit Sales Report"

    # Write headers
    headers = ["", "Reference No", "Customer Name", "Building Name", "Net taxable", "Vat","Grand Total"]
    worksheet.write_row(0, 0, headers)

    # Write data rows
    for index, invoice in enumerate(invoices, start=1):
        worksheet.write_row(index, 0, [
            index,
            invoice.reference_no,
            invoice.customer.customer_name,
            invoice.customer.building_name,
            invoice.net_taxable,
            invoice.vat,
            invoice.amout_total
        ])
    #  'salesman_customers_count': salesman_customers_count,
    #     'emergency_customers': emergency_customers,
    #     'visited_customers_count': visited_customers_count,
    #     'non_visited_customers': non_visited_customers,
    #     'planned_visits_count': planned_visits_count

    # # Write footer data
    # footer_row = len(invoices) + 1
    # worksheet.write(footer_row, 3, "Total")
    # worksheet.write(footer_row, 4, total_net_taxable)
    # worksheet.write(footer_row, 5, total_vat)
    # worksheet.write(footer_row, 6, total_amout_total)

    # Close the workbook
    workbook.close()

    # Create HttpResponse object to return the Excel file as a response
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Visit_Statistics_Report.xlsx'
    return response


def visitstatistics_report_print(request):
    user_id = request.user.id

    start_date = request.GET.get('start_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
    
    salesman_customers_count = Customers.objects.filter(created_date__date=start_date,sales_staff_id=user_id).count()
    #emergency supply
    emergency_customers = DiffBottlesModel.objects.filter(created_date__date=start_date, assign_this_to_id=user_id).count()
    print('emergency_customers',emergency_customers)
    #actual visit
    visited_customers_count = CustomerSupply.objects.filter(salesman_id=user_id, created_date__date=start_date).count()
    print(visited_customers_count,'visited_customers_count')
    
    planned_visits_count = Customers.objects.filter(visit_schedule__isnull=False, created_date__date=start_date).annotate(
        planned_visit_count=Count('visit_schedule')
    ).filter(planned_visit_count__gt=0).count()
    print(planned_visits_count,'planned_visits_count')

    # non_visited_customers = CustomerSupply.objects.exclude(salesman_id=user_id, created_date__date=start_date).count()
    non_visited_customers = planned_visits_count - visited_customers_count

   


    # Create a BytesIO object to save PDF data
    buffer = BytesIO()

    # Create a new PDF document
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    
    # Define table data
    data = [
        ["Sl No", "Reference No", "Customer Name", "Building Name", "Net taxable", "Vat","Grand Total"]
    ]
    
    # Populate table data
    for index, invoice in enumerate(invoices, start=1):
        data.append([
            str(index),
            str(invoice.reference_no),
            str(invoice.customer.customer_name),
            str(invoice.customer.building_name),
            str(invoice.net_taxable),
            str(invoice.vat),
            str(invoice.amout_total)
        ])
    
   

    # Create a table
    table = Table(data)

    # Add style to the table
    style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ])
    table.setStyle(style)

    # Add table to the document
    doc.build([table])

    # Create HttpResponse object to return the PDF file as a response
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=Visit_Statistics_Report.pdf'
    return response


def fivegallonrelated_report(request):
    user_id = request.user.id
    start_date = request.GET.get('start_date')
    filter_data = {}

    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()

    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')
    empty_bottles_collected = CustomerSupply.objects.filter(created_date__date=start_date, collected_empty_bottle__gt=0, salesman_id=user_id).count()

    # empty_bottles_collected = CustomerSupply.objects.filter(created_date__date=start_date, collected_empty_bottle__gt=0, **filter_data).count()
    empty_bottle_pending = CustomerSupply.objects.filter(created_date__date=start_date, allocate_bottle_to_pending__gt=0, salesman_id=user_id).count()
    print(empty_bottle_pending,'empty_bottle_pending')
    coupons_collected = CustomerSupplyCoupon.objects.filter(customer_supply__created_date__date=start_date, customer_supply__salesman_id=user_id).aggregate(total_coupons=Count('leaf'))['total_coupons']
    total_supplied_quantity = CustomerSupplyItems.objects.filter(customer_supply__created_date__date=start_date, customer_supply__salesman_id=user_id).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
    total_collected_amount = CustomerSupply.objects.filter(created_date__date=start_date, salesman_id=user_id).aggregate(total_collected_amount=Sum('net_payable'))['total_collected_amount'] or 0
    total_pending_amount = CustomerSupply.objects.filter(created_date__date=start_date, salesman_id=user_id).aggregate(total_pending_amount=Sum('grand_total') - Sum('net_payable'))['total_pending_amount'] or 0
    mode_of_supply = CustomerSupply.objects.filter(created_date__date=start_date, salesman_id=user_id).values('customer__sales_type').annotate(total=Count('customer__sales_type'))

    context = {
        'empty_bottles_collected': empty_bottles_collected,
        'empty_bottle_pending': empty_bottle_pending,
        'coupons_collected': coupons_collected,
        'total_supplied_quantity': total_supplied_quantity,
        'total_collected_amount': total_collected_amount,
        'total_pending_amount': total_pending_amount,
        'mode_of_supply' :mode_of_supply
    }

    return render(request, 'sales_management/dsr_fivegallonrelated_report.html', context)

    
def bottlecount_report(request):
    user_id = request.user.id
    start_date = request.GET.get('start_date')
    filter_data = {}
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()

    filter_data['start_date'] = start_date.strftime('%Y-%m-%d')
    
    total_empty_bottles = CustomerCustodyStock.objects.filter(customer__sales_staff_id=user_id).aggregate(total_quantity=Sum('quantity'))['total_quantity']
    if total_empty_bottles is None:
        total_empty_bottles = 0

    total_supplied_bottles = CustomerSupply.objects.filter(created_date__date=start_date).aggregate(total_bottles=Sum('collected_empty_bottle'))['total_bottles']
    # closing_stock_count = VanStock.objects.filter(created_date=start_date,stock_type="closing").aggregate(total_count=Sum('count'))['total_count'] or 0,
    closing_stock_count = VanStock.objects.filter(created_date=start_date, stock_type='closing').count() or 0
    damage_bottle_count = VanProductItems.objects.filter(van_stock__created_date=start_date, van_stock__stock_type='damage').aggregate(total_damage=Sum('count'))['total_damage'] or 0
    pending_bottle_count = CustomerSupply.objects.filter(created_date__date=start_date,salesman_id=user_id).aggregate(total_pending=Sum('allocate_bottle_to_pending'))['total_pending'] or 0
    if pending_bottle_count is None:
        pending_bottle_count = 0

    total_count = total_empty_bottles + total_supplied_bottles + closing_stock_count + damage_bottle_count + pending_bottle_count

    context = {
        'total_empty_bottles': total_empty_bottles,
        'total_supplied_bottles':total_supplied_bottles,
        'closing_stock_count': closing_stock_count,
        'damage_bottle_count': damage_bottle_count,
        'pending_bottle_count': pending_bottle_count,
        'total_count': total_count,
    }

    return render(request, 'sales_management/dsr_bottlecount_report.html', context)

