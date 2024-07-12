import re
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from coupon_management.serializers import couponStockSerializers
from lxml.etree import HTML
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate

from client_management.models import *
from competitor_analysis.forms import CompetitorAnalysisFilterForm
from master.functions import generate_form_errors
from product.models import Staff_Orders_details
from .models import *
from .forms import  *
from accounts.models import CustomUser, Customers
from master.models import EmirateMaster, BranchMaster, RouteMaster
import json
from django.core.serializers import serialize
from django.views import View
from datetime import datetime
from django.core.exceptions import ObjectDoesNotExist
from django.http import JsonResponse, HttpResponse

# Create your views here.
def get_next_coupon_bookno(request):
    coupon_type = request.GET.get("coupon_type")
    last_coupon = NewCoupon.objects.filter(coupon_type__pk=coupon_type).latest("created_date")
    last_coupon_bookno = last_coupon.book_num

    # Split the alphanumeric book number into alphabetic and numeric parts
    match = re.match(r"([a-zA-Z]+)(\d+)", last_coupon_bookno)
    if not match:
        raise ValueError("Invalid book number format")

    alphabetic_part, numeric_part = match.groups()
    next_numeric_part = int(numeric_part) + 1
    next_coupon_bookno = f"{alphabetic_part}{next_numeric_part}"

    data = {'next_coupon_bookno': next_coupon_bookno}
    return JsonResponse(data, safe=False)


def get_coupon_bookno(request):
    request_id = request.GET.get("request_id")
    
    if (instances:=Staff_Orders_details.objects.filter(pk=request_id)).exists():
        instance = instances.first()
        stock_instances = CouponStock.objects.filter(couponbook__coupon_type__coupon_type_name=instance.product_id.product_name,coupon_stock="company")
        serialized = couponStockSerializers(stock_instances, many=True)
        
        status_code = 200
        response_data = {
            "status": "true",
            "data": serialized.data,
        }
    else:
        status_code = 404
        response_data = {
            "status": "false",
            "title": "Failed",
            "message": "item not found",
        }

    return HttpResponse(json.dumps(response_data),status=status_code, content_type="application/json")

def couponType(request):
    all_couponType = CouponType.objects.all()
    context = {'all_couponType': all_couponType}
    return render(request, 'coupon_management/index_couponType.html', context)

def create_couponType(request):
    if request.method == 'POST':
        form = CreateCouponTypeForm(request.POST)
        if form.is_valid():
            data = form.save(commit=False)
            data.created_by = str(request.user.id)
            data.save()
            messages.success(request, 'Coupon Type created successfully!')
            return redirect('couponType')
        else:
            messages.error(request, 'Invalid form data. Please check the input.')
    else:
        form = CreateCouponTypeForm()
    context = {'form': form}
    return render(request, 'coupon_management/create_couponType.html', context)


def view_couponType(request, coupon_type_id):
    view_couponType = get_object_or_404(CouponType, coupon_type_id=coupon_type_id)
    return render(request, 'coupon_management/view_couponType.html', {'view_couponType': view_couponType})

def edit_CouponType(request, coupon_type_id):
    edit_coupon = get_object_or_404(CouponType, coupon_type_id=coupon_type_id)
    if request.method == 'POST':
        form = EditCouponTypeForm(request.POST, instance=edit_coupon)
        if form.is_valid():
            data = form.save(commit=False)
            # print(data,"data")
            data.modified_by = str(request.user.id)
            data.modified_date = datetime.now()
            data.save()
            return redirect('couponType')
    else:
        form = EditCouponTypeForm(instance=edit_coupon)
    return render(request, 'coupon_management/edit_couponType.html', {'form': form, 'edit_coupon': edit_coupon})

def delete_couponType(request, coupon_type_id):
    deleteCouponType = CouponType.objects.get(coupon_type_id=coupon_type_id)
    if request.method == 'POST':
        deleteCouponType.delete()
        return redirect('couponType')
    return render(request, 'coupon_management/delete_couponType.html', {'deleteCouponType': deleteCouponType})

#------------------------New Coupon
def new_coupon(request):
    all_coupon = NewCoupon.objects.all().order_by("-created_date")
    coupon_stock_list = CouponStock.objects.all()  # Fetch all CouponStock instances
    context = {'all_coupon': all_coupon, 'coupon_stock_list': coupon_stock_list}  # Add coupon_stock_list to context
    return render(request, 'coupon_management/index_Newcoupon.html', context)

def create_Newcoupon(request):
    if request.method == 'POST':
        form = CreateNewCouponForm(request.POST)
        if form.is_valid():
            data = form.save(commit=False)
            
            coupon_type_id = request.POST.get('coupon_type')
            book_num = request.POST.get('book_num')
            
            # Check if the coupon book already exists
            if NewCoupon.objects.filter(book_num=book_num,coupon_type=coupon_type_id).exists():
                # print("Exists")
                messages.error(request, f'Coupon book {book_num} already exists.')
                return redirect('new_coupon')
            else:
                selected_coupon_type = get_object_or_404(CouponType, coupon_type_id=coupon_type_id)

                data.coupon_type = selected_coupon_type
                data.book_num = book_num
                data.no_of_leaflets = selected_coupon_type.no_of_leaflets
                data.valuable_leaflets = selected_coupon_type.valuable_leaflets
                data.free_leaflets = selected_coupon_type.free_leaflets
                data.created_by = str(request.user.id)
                branch_id = request.user.branch_id.branch_id
                branch = BranchMaster.objects.get(branch_id=branch_id)
                data.branch_id = branch           
                data.save()
                print("DATA SAVED")

                # Generate leaflets for the coupon
                leaflets = []
                no_of_leaflets = int(selected_coupon_type.no_of_leaflets)
                print("no_of_leaflets",no_of_leaflets)
                for leaflet_num in range(1, no_of_leaflets + 1):
                    if leaflet_num < 10:  # If leaflet number is less than 10, add a leading '0'
                        leaflet_name = f"{book_num}0{leaflet_num}"
                    else:
                        leaflet_name = f"{book_num}{leaflet_num}"  # Otherwise, use leaflet_num directly
                    print("leaflet_name", leaflet_name)

                    leaflet = CouponLeaflet(coupon=data, leaflet_number=str(leaflet_num),leaflet_name=leaflet_name)
                    print("leaflet",leaflet)
                    leaflet.save()
                    leaflets.append({'leaflet_number': leaflet.leaflet_number})
                # Create CouponStock instance
                coupon_stock = CouponStock.objects.create(couponbook=data, coupon_stock='company', created_by=str(request.user.id))
                # print("coupon_stock",coupon_stock)

                response_data = {
                    'success': True,
                    'book_num': data.book_num,
                    'leaflets': leaflets
                }
                return JsonResponse(response_data, status=200)
        else:
            message = generate_form_errors(form,formset=False)
            response_data = {
                'success': False, 
                'message': message,
                }
            return JsonResponse(response_data, status=200)
    else:
        form = CreateNewCouponForm()
    
    context = {'form': form}
    return render(request, 'coupon_management/create_Newcoupon.html', context)

def generate_leaflets(request, coupon_id):
    coupon = get_object_or_404(NewCoupon, coupon_id=coupon_id)
    leaflets = []
    no_of_leaflets = int(coupon.coupon_type.no_of_leaflets)
    for leaflet_num in range(1, no_of_leaflets + 1):
        leaflet = CouponLeaflet(coupon=coupon, leaflet_number=str(leaflet_num))
        print("leaflet",leaflet)
        leaflets.append(leaflet)
        leaflet.save()
        

    context = {'coupon': coupon, 'leaflets': leaflets}
    return render(request, 'coupon_management/create_Newcoupon.html', context)

def get_leaflet_serial_numbers(request):
    if request.method == 'GET':
        coupon_type_id = request.GET.get('coupon_type')
        print("coupon_type_id",coupon_type_id)

        # Fetch leaflets based on the provided coupon type
        try:
            leaflets = CouponLeaflet.objects.filter(coupon__coupon_type_id=coupon_type_id)
            print("leaflets",leaflets)
            leaflet_data = [{'leaflet_number': leaflet.leaflet_number, 'is_used': leaflet.used} for leaflet in leaflets]
            print("leaflet_data",leaflet_data)
            return JsonResponse(leaflet_data, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    

def save_coupon_data(request):
    if request.method == 'POST':
        form = CreateNewCouponForm(request.POST)
        if form.is_valid():
            try:
                # Save the coupon data to the database
                new_coupon = form.save()
                # You can also save leaflet data here if necessary
                return JsonResponse({'message': 'Coupon data saved successfully'})
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=500)
        else:
            return JsonResponse({'error': 'Invalid form data'}, status=400)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)

def view_Newcoupon(request, coupon_id):
    view_coupon = get_object_or_404(NewCoupon, coupon_id=coupon_id)
    return render(request, 'coupon_management/view_Newcoupon.html', {'view_coupon': view_coupon})

def edit_NewCoupon(request, coupon_id):
    edit_coupon = get_object_or_404(NewCoupon, coupon_id=coupon_id)
    if request.method == 'POST':
        form = EditNewCouponForm(request.POST, instance=edit_coupon)
        if form.is_valid():
            data = form.save(commit=False)
            # print(data,"data")
            data.modified_by = str(request.user.id)
            data.modified_date = datetime.now()
            data.save()
            return redirect('new_coupon')
    else:
        form = EditNewCouponForm(instance=edit_coupon)
    return render(request, 'coupon_management/edit_Newcoupon.html', {'form': form, 'edit_coupon': edit_coupon})

def delete_Newcoupon(request, coupon_id):
    deleteCoupon = NewCoupon.objects.get(coupon_id=coupon_id)
    if request.method == 'POST':
        deleteCoupon.delete()
        return redirect('new_coupon')
    return render(request, 'coupon_management/delete_Newcoupon.html', {'deleteCoupon': deleteCoupon})


# def customer_stock(request):
#
#     couponstock=CustomerCouponStock.objects.select_related('customer').all()
#     context = {
#
#         'couponstock':couponstock
#     }
#
#     return render(request, 'coupon_management/customer_stock.html', context)


# def customer_stock(request):
#     coupenstock = CustomerCouponStock.objects.select_related('customer').all()
#     return render(request, 'coupon_management/customer_stock.html', {'coupenstock': coupenstock})



#
# @login_required
# def customer_stock(request):
#     # Get the user's user_type
#     user_type = request.user.user_type
#
#     # Filter customers based on user_type
#     if user_type == 'Salesman':
#         coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__sales_staff=request.user)
#         route_li = RouteMaster.objects.filter(branch_id=request.user.branch_id)
#     else:
#         coupenstock = CustomerCouponStock.objects.select_related('customer').all()
#         route_li = RouteMaster.objects.all()
#
#     return render(request, 'coupon_management/customer_stock.html', {'coupenstock': coupenstock, 'route_li': route_li})



@login_required
def customer_stock(request):
    # Get the user's user_type
    user_type = request.user.user_type

    # Get all routes for the dropdown
    route_li = RouteMaster.objects.all()

    # Filter customers based on user_type and selected route
    selected_route = request.GET.get('route_name')
    if user_type == 'Salesman':
        if selected_route:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__sales_staff=request.user, customer__routes__route_name=selected_route)
        else:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__sales_staff=request.user)
    else:
        if selected_route:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__routes__route_name=selected_route)
        else:
            coupenstock = CustomerCouponStock.objects.select_related('customer').all()


    created_date = request.GET.get('created_date')
    if created_date:
        coupenstock = coupenstock.filter(
            coupon_type_id__created_date=created_date
        )

    return render(request, 'coupon_management/customer_stock.html', {'coupenstock': coupenstock, 'route_li': route_li})







from openpyxl.styles import Font, Alignment

@login_required
def generate_excel(request):
    # Get the user's user_type
    user_type = request.user.user_type

    # Get all routes for the dropdown
    route_li = RouteMaster.objects.all()

    # Filter customers based on user_type and selected route
    selected_route = request.GET.get('route_name')
    if user_type == 'Salesman':
        if selected_route:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__sales_staff=request.user, customer__routes__route_name=selected_route)
        else:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__sales_staff=request.user)
    else:
        if selected_route:
            coupenstock = CustomerCouponStock.objects.select_related('customer').filter(customer__routes__route_name=selected_route)
        else:
            coupenstock = CustomerCouponStock.objects.select_related('customer').all()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="customer_stock.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Customer Stock'

    # Formatting styles
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    data_alignment = Alignment(horizontal='left', vertical='center')

    # Write headers with formatting
    headers = [
        'Sl.No', 'Customer Name / Mobile No', 'Building Name / House No',
        'Digital Coupon Count', 'Manual Coupon Count', 'Total Count'
    ]
    for col_num, header_title in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data with formatting
    for row_num, stock in enumerate(coupenstock, start=2):
        worksheet.cell(row=row_num, column=1, value=row_num - 1).alignment = data_alignment
        worksheet.cell(row=row_num, column=2, value=stock.customer.customer_name + ', ' + stock.customer.mobile_no).alignment = data_alignment
        worksheet.cell(row=row_num, column=3, value=stock.customer.building_name + ', ' + stock.customer.door_house_no).alignment = data_alignment
        if stock.coupon_method == 'digital':
            worksheet.cell(row=row_num, column=4, value=stock.count).alignment = data_alignment
        else:
            worksheet.cell(row=row_num, column=5, value=stock.count).alignment = data_alignment

    # Calculate and write total counts
    total_digital_count = sum(stock.count for stock in coupenstock if stock.coupon_method == 'digital')
    total_manual_count = sum(stock.count for stock in coupenstock if stock.coupon_method == 'manual')
    total_row_num = len(coupenstock) + 2
    worksheet.cell(row=total_row_num, column=4, value=total_digital_count).alignment = data_alignment
    worksheet.cell(row=total_row_num, column=5, value=total_manual_count).alignment = data_alignment
    worksheet.cell(row=total_row_num, column=6, value=total_digital_count + total_manual_count).alignment = data_alignment

    # Autofit column width
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)
    return response

from io import BytesIO
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors


def customer_stock_pdf(request):
    # Retrieve customer stock data
    coupenstock = CustomerCouponStock.objects.select_related('customer').all()

    # Check if coupenstock is not empty
    if coupenstock:
        # Create the HTTP response with PDF content type and attachment filename
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="customer_stock.pdf"'

        # Create a PDF document
        pdf_buffer = BytesIO()
        pdf = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        data = []

        # Add headers
        headers = ['Sl.No', 'Customer Name / Mobile No', 'Building Name / House No', 'Digital', 'Manual', 'Total Count']
        data.append(headers)

        # Add data to the PDF document
        for index, stock in enumerate(coupenstock):
            customer_name_mobile = f"{stock.customer.customer_name}, {stock.customer.mobile_no}"
            building_house = f"{stock.customer.building_name}, {stock.customer.door_house_no}"
            digital = stock.count if stock.coupon_method == 'digital' else ''
            manual = stock.count if stock.coupon_method == 'manual' else ''
            total_count = stock.count
            data.append([index + 1, customer_name_mobile, building_house, digital, manual, total_count])

        table = Table(data)
        style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)])
        table.setStyle(style)

        # Add the table to the PDF document
        elements = [table]
        pdf.build(elements)

        # Get the value of the BytesIO buffer and write it to the response
        pdf_value = pdf_buffer.getvalue()
        pdf_buffer.close()
        response.write(pdf_value)

        return response
    else:
        # Return an empty HTTP response with a message indicating no data available
        return HttpResponse('No customer stock data available.')
    





# def get_supplied_manual_coupons_count():
    
#     supplied_coupons = CustomerSupplyCoupon.objects.annotate(
#         manual_coupons_count=Count('leaf__coupon__coupon_type', filter=Q(leaf__coupon__coupon_method='manual'))
#     ).values('customer_supply').annotate(total_manual_coupons=Sum('manual_coupons_count'))

   
#     supplied_coupons_count = {supply['customer_supply']: supply['total_manual_coupons'] for supply in supplied_coupons}

#     return supplied_coupons_count

# def redeemed_history(request):
#     instances = CustomerSupplyCoupon.objects.all().order_by("-created_date")

#     start_date_str = request.GET.get('start_date')
#     end_date_str = request.GET.get('end_date')
    
#     if start_date_str and end_date_str:
#         start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
#         end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
#         instances = instances.filter(created_date__range=[start_date, end_date])
    
#     supplied_coupons_count = get_supplied_manual_coupons_count()
#     print(supplied_coupons_count,"supplied_coupons_count")

#     context = {
#         'instances': instances,
#         'supplied_coupons_count': supplied_coupons_count
#     }

#     return render(request, 'client_management/redeemed_history.html', context)
from django.db.models import Count, Q

def get_customer_coupon_counts():
    
    customer_coupon_counts = CustomerSupplyCoupon.objects.values('customer_supply__customer').annotate(
        manual_coupons_count=Count('leaf__coupon__coupon_type', filter=Q(leaf__coupon__coupon_method='manual')),
        digital_coupons_count=Count('leaf__coupon__coupon_type', filter=Q(leaf__coupon__coupon_method='digital'))
    )
    manual_coupons_count=Count('leaf__coupon__coupon_type', filter=Q(leaf__coupon__coupon_method='manual'))

    digital_coupons_count=Count('leaf__coupon__coupon_type', filter=Q(leaf__coupon__coupon_method='digital'))
    print("customer_coupon_counts",customer_coupon_counts)

    customer_coupons = {}
    for coupon_count in customer_coupon_counts:
        customer_id = coupon_count['customer_supply__customer']
        customer_coupons[customer_id] = {
            'manual_coupons': coupon_count['manual_coupons_count'],
            'digital_coupons': coupon_count['digital_coupons_count']
        }

    return customer_coupons

def redeemed_history(request):
    
    instances = CustomerSupply.objects.all().order_by("-created_date")

    instances_coupon = CustomerSupplyCoupon.objects.filter(customer_supply__in=instances).order_by("-customer_supply__created_date")

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        instances_coupon = instances_coupon.filter(customer_supply__created_date__range=[start_date, end_date])
    
    customer_coupon_counts = get_customer_coupon_counts()
    print(customer_coupon_counts,'customer_coupon_counts')

    context = {
        'instances': instances,
        'instances_coupon': instances_coupon,  
        'customer_coupon_counts': customer_coupon_counts
    }

    return render(request, 'coupon_management/redeemed_history.html', context)
