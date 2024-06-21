from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import *
# from .serializers import CustomerOrderSerializer
from django.shortcuts import render, redirect
from django.views import View
from .forms import *
from product.models import Product
from master.models import RouteMaster
from django.utils import timezone
import datetime
import io
import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from .forms import OrderForm
from django.shortcuts import get_object_or_404
from datetime import date as dt

# class CustomerOrder(APIView):
#     def post(self, request, *args, **kwargs):
#         serializer = CustomerOrderSerializer(data=request.data)
#         print(serializer,"serializer")
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_201_CREATED)
        # return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


def order_change_reason(request):
    template = 'order/reason_list.html'
    reasons = Change_Reason.objects.all()
    context = {'reasons':reasons}
    return render(request, template, context)


    
class Reason_Add(View):
    def get(self, request):
        template = 'order/reason_add.html'
        form = Reason_Add_Form
        
        return render(request, template, {'form': form})

    def post(self, request):
        template = 'order/reason_add.html'
        form = Reason_Add_Form(request.POST)
        if form.is_valid():
            form.save()
            return redirect(order_change_reason)
        return render(request, template, {'form': form})
    
class Reason_Edit(View):
    template = 'order/reason_edit.html'
    def get(self, request, reason_id):
        reason = Change_Reason.objects.get(id=reason_id)
        form = Reason_Edit_Form(instance=reason)
        return render(request, self.template, {'form': form, 'reason': reason})

    def post(self, request, reason_id):
        reason = Change_Reason.objects.get(id=reason_id)
        form = Reason_Edit_Form(request.POST, instance=reason)
        if form.is_valid():
            form.save()
            return redirect(order_change_reason)
        return render(request, self.template, {'form': form, 'reason': reason})

class Reason_Delete(View):
    template='order/reason_delete.html'
    def get(self, request, reason_id):
        reason = Change_Reason.objects.get(id=reason_id)
        return render(request, self.template, {'reason': reason})

    def post(self, request, reason_id):
        reason = Change_Reason.objects.get(id=reason_id)
        reason.delete()
        return redirect(order_change_reason)
    

from django.utils import timezone
# Order_change


def filter_route(request, action):
    routes = RouteMaster.objects.all()
    if request.method == 'POST':
        route = request.POST.get('route')
        if action == 'change':
            return redirect('order_change_add', route_id = route)
        if action == 'return':
            return redirect('order_return_add', route_id = route)
    return render(request, 'order/filter_route.html', {'routes':routes, 'action':action})

    

# Change
def order_change(request):
    start_date = None
    end_date = None
    selected_date = None
    selected_product_id = None
    selected_product = None
    template = 'order/order_change.html'
    order_changes = Order_change.objects.all()
    products = ProdutItemMaster.objects.all()
    route_counts = {}
    today = datetime.today()
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        selected_date = request.POST.get('date')
        selected_product_id = request.POST.get('product')
        if start_date and end_date:
            order_changes = order_changes.filter(change_date__range=[start_date, end_date])
        elif selected_date:
            order_changes = order_changes.filter(change_date=selected_date)
        else:
            order_changes=Order_change.objects.filter(change_date=timezone.now().date())
        
        if selected_product_id:
            selected_product = Product.objects.get(product_id = selected_product_id)
            order_changes = order_changes.filter(product=selected_product)
    else:
        order_changes=Order_change.objects.filter(change_date=timezone.now().date())
    
    for order_change in order_changes:
        
        route = order_change.route
        if route:
            route_name = route.route_name
            route_id = route.route_id  
            q=order_change.changed_quantity
            if route_name not in route_counts:
                route_counts[route_name] = {'count': q, 'id': route_id}  
            else:
                route_counts[route_name]['count'] += q
        

    context = {'order_changes':order_changes, 'products' :products, 'route_counts': route_counts, 'today':today,
               'start_date':start_date, 'end_date':end_date, 'selected_date':selected_date, 'selected_product_id':selected_product_id, 'selected_product':selected_product}
    return render(request, template, context)

import datetime
from datetime import date as dt
def order_change_list(request, route_id):
    route = RouteMaster.objects.get(route_id=route_id)
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    selected_date = request.GET.get('selected_date')
    selected_product_id = request.GET.get('selected_product_id', None)
    # date = ''
    today = datetime.today()
    order_changes = Order_change.objects.filter(customer__routes=route)
   
    if start_date and end_date and start_date != 'None' and end_date != 'None':
        order_changes = order_changes.filter(change_date__range=[start_date, end_date])
        date=f" {start_date} --> { end_date }"
    elif selected_date and selected_date != 'None':
        order_changes = order_changes.filter(change_date=selected_date)
        date=f"{selected_date}"
    else:
        today = dt.today()
        order_changes = order_changes.filter(change_date=today)
        date=f"{today}"
    if selected_product_id != 'None':
        selected_product=Product.objects.get(product_id=selected_product_id)
        order_changes = order_changes.filter(product=selected_product)
    else:
        selected_product=None
    # Render the template with necessary context
    
    
    context = {
        'order_changes': order_changes,
        'route':route,
        'start_date': start_date,
        'end_date': end_date,
        'selected_date': selected_date,
        
        'selected_product': selected_product,
        'date':date,

    }
    return render(request, 'order/order_change_list.html', context)



def order_change_list_excel(request, route_id):
    route = RouteMaster.objects.get(route_id=route_id)
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    selected_date = request.GET.get('selected_date')
    selected_product_id = request.GET.get('selected_product_id', None)
    # date = ''
    today = datetime.today()
    order_changes = Order_change.objects.filter(customer__routes=route)
   
    if start_date and end_date and start_date != 'None' and end_date != 'None':
        order_changes = order_changes.filter(change_date__range=[start_date, end_date])
        date=f" {start_date} --> { end_date }"
    elif selected_date and selected_date != 'None':
        order_changes = order_changes.filter(change_date=selected_date)
        date=f"{selected_date}"
    else:
        today = dt.today()
        order_changes = order_changes.filter(change_date=today)
        date=f"{today}"
    if selected_product_id and selected_product_id != 'None':
        selected_product=Product.objects.get(product_id=selected_product_id)
        order_changes = order_changes.filter(product=selected_product)
    else:
        selected_product=None

    data = []
    data = list(order_changes.values_list('customer__customer_name', 'customer__mobile_no', 
                                          'customer__location__location_name', 
                                          'customer__building_name', 'customer__floor_no', 'customer__door_house_no', 
                                          'product__product_name__product_name', 
                                          'changed_quantity', 'reason__reason_name'
                                        ))
    df = pd.DataFrame(data, columns=['Customer ', 'Mobile no', 'Location', ' Building', 'Floor no', 'Door no', 'Product Name', 'Quantity ', 'Reason'])

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Add information above the table by merging cells
    
    ws.merge_cells('A1:J2')  # Merge cells for the title
    ws['A1'] = 'Sana Water'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A3:J3')
    ws['A3'] = 'Changed products Information' 
    ws['A3'].alignment = Alignment(horizontal='center') 
    ws.merge_cells('A4:B4')
    ws['A4'] = f'Route: {route.route_name}' 
    ws.merge_cells('C4:F4') 
    ws['C4'] = f'Date: {date}' 
    ws.merge_cells('G4:J4') 
    if selected_product:
        ws['G4'] = f'Selected Product: {selected_product}'  
    headers =['SiNo']+ ['Customer ', 'Mobile no', 'Location', ' Building', 'Floor no', 'Door no', 'Product Name', 'Quantity ', 'Reason']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.font = Font(bold=True) 
    # Write DataFrame to the Excel file starting from row 6
    for r_idx, row in enumerate(df.values, start=7):  # Start from index 7
        ws.cell(row=r_idx, column=1, value=r_idx - 6)  # Add serial number starting from 1
        for c_idx, value in enumerate(row, start=2):  # Start from column 2 to avoid overwriting serial number
            ws.cell(row=r_idx, column=c_idx, value=value)
    # Save the workbook to a BytesIO object
    from io import BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Serve the Excel file as an HTTP response
    response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=changed-products.xlsx'
    return response

    
    

class Order_change_Add(View):
    def get(self, request, route_id):
        template = 'order/order_change_add.html'
        form = Order_change_Form(route_id=route_id)
        
        return render(request, template, {'form': form})

    def post(self, request, route_id):
        template = 'order/order_change_add.html'
        form = Order_change_Form(route_id=route_id, data=request.POST)
        if form.is_valid():
            route = RouteMaster.objects.get(route_id=route_id)
            form.save(route=route)
            return redirect(order_change)
        customer = request.POST.get('customer')
        return render(request, template,)


class Order_change_Edit(View):
    template = 'order/order_change_edit.html'
    def get(self, request, order_change_id):
        customer = Order_change.objects.get(order_change_id=order_change_id).customer
        order_change = Order_change.objects.get(order_change_id=order_change_id)
        form = Order_change_Edit_Form(instance=order_change)
        return render(request, self.template, {'form': form, 'order_change': order_change, 'customer':customer})

    def post(self, request, order_change_id):
        order_change = Order_change.objects.get(order_change_id=order_change_id)
        form =Order_change_Edit_Form(request.POST, instance=order_change)
        
        if form.is_valid():
            form.save()
            return redirect('order_change')
        else:
            print(form.errors)
        return render(request, self.template, {'form': form, 'order_change': order_change})
    
class Order_change_Delete(View):
    template='order/order_change_delete.html'
    def get(self, request, order_change_id):
        order_change = Order_change.objects.get(order_change_id=order_change_id)
        return render(request, self.template, {'order_change': order_change})

    def post(self, request, order_change_id):
        order_change = Order_change.objects.get(order_change_id=order_change_id)
        order_change.delete()
        return redirect('order_change')


# Return
def order_return(request):
    start_date = None
    end_date = None
    selected_date = None
    selected_product_id = None
    selected_product = None
    template = 'order/order_return.html'
    order_returns = Order_return.objects.all()
    products = ProdutItemMaster.objects.all()
    route_counts = {}
    today = datetime.today()
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        selected_date = request.POST.get('date')
        selected_product_id = request.POST.get('product')

        if start_date and end_date:
            order_returns = order_returns.filter(return_date__range=[start_date, end_date])
        elif selected_date:
            order_returns = order_returns.filter(return_date=selected_date)
        else:
            order_returns=Order_return.objects.filter(return_date=timezone.now().date())
        
            
        
        if selected_product_id:
            selected_product = Product.objects.get(product_id = selected_product_id)
            order_returns = order_returns.filter(product=selected_product)
    else:
        order_returns=Order_return.objects.filter(return_date=timezone.now().date())
    
    for order_return in order_returns:
        
        route = order_return.route
        if route:
            route_name = route.route_name
            route_id = route.route_id  # Include the route_id
            q=order_return.returned_quantity
            if route_name not in route_counts:
                route_counts[route_name] = {'count': q, 'id': route_id}  
            else:
                route_counts[route_name]['count'] += q
        

    context = {'order_returns':order_returns, 'products' :products, 'route_counts': route_counts, 'today':today,
               'start_date':start_date, 'end_date':end_date, 'selected_date':selected_date, 'selected_product_id':selected_product_id, 'selected_product':selected_product}
    return render(request, template, context)

import datetime
from datetime import date as dt
def order_return_list(request, route_id):
    # Retrieve route object
    route = RouteMaster.objects.get(route_id=route_id)
    
    # Get query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    selected_date = request.GET.get('selected_date')
    selected_product_id = request.GET.get('selected_product_id', None)
    # date = ''
    order_returns = Order_return.objects.filter(customer__routes=route)
    
    if start_date and end_date and start_date != 'None' and end_date != 'None':
        order_returns = order_returns.filter(return_date__range=[start_date, end_date])
        date=f" {start_date} --> { end_date }"
    elif selected_date and selected_date != 'None':
        order_returns = order_returns.filter(return_date=selected_date)
        date=f"{selected_date}"
    else:
        today = dt.today()
        order_returns = order_returns.filter(return_date=today)
        date=f"{today}"
    if selected_product_id != 'None':
        
        selected_product=Product.objects.get(product_id=selected_product_id)
        order_returns = order_returns.filter(product=selected_product)
    else:
        selected_product=None
    # Render the template with necessary context
    
    
    context = {
        'order_returns': order_returns,
        'route':route,
        'start_date': start_date,
        'end_date': end_date,
        'selected_date': selected_date,
        'selected_product_id':selected_product_id,
        'selected_product': selected_product,
        'date':date
    }
    return render(request, 'order/order_return_list.html', context)



def order_return_list_excel(request, route_id):
    route = RouteMaster.objects.get(route_id=route_id)
    
    # Get query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    selected_date = request.GET.get('selected_date')
    selected_product_id = request.GET.get('selected_product_id', None)
    # date = ''
    order_returns = Order_return.objects.filter(customer__routes=route)
    if start_date and end_date and start_date != 'None' and end_date != 'None':
        order_returns = order_returns.filter(return_date__range=[start_date, end_date])
        date=f" {start_date} --> { end_date }"
    elif selected_date and selected_date != 'None':
        order_returns = order_returns.filter(return_date=selected_date)
        date=f"{selected_date}"
    else:
        today = dt.today()
        order_returns = order_returns.filter(return_date=today)
        date=f"{today}"
    if selected_product_id != 'None':
        
        selected_product=Product.objects.get(product_id=selected_product_id)
        order_returns = order_returns.filter(product=selected_product)
    else:
        selected_product=None
    data = []
    data = list(order_returns.values_list('customer__customer_name', 'customer__mobile_no', 
                                          'customer__location__location_name', 
                                          'customer__building_name', 'customer__floor_no', 'customer__door_house_no', 
                                          'product__product_name__product_name', 
                                          'returned_quantity', 'reason__reason_name'
                                        ))
    df = pd.DataFrame(data, columns=['Customer ', 'Mobile no', 'Location', ' Building', 'Floor no', 'Door no',  'Product Name', 'Quantity ', 'Reason'])
    print(data)
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Add information above the table by merging cells
    # Add information above the table by merging cells
    
    ws.merge_cells('A1:J2')  # Merge cells for the title
    ws['A1'] = 'Sana Water'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A3:J3')
    ws['A3'] = 'Returned products Information' 
    ws['A3'].alignment = Alignment(horizontal='center') 
    ws.merge_cells('A4:B4')
    ws['A4'] = f'Route: {route.route_name}' 
    ws.merge_cells('C4:F4') 
    ws['C4'] = f'Date: {date}' 
    ws.merge_cells('G4:J4') 
    if selected_product:
        ws['G4'] = f'Selected Product: {selected_product}'  
    headers =['SiNo']+ ['Customer ', 'Mobile no', 'Location', ' Building', 'Floor no', 'Door no', 'Product Name', 'Quantity ', 'Reason']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.font = Font(bold=True) 
    # Write DataFrame to the Excel file starting from row 6
    for r_idx, row in enumerate(df.values, start=7):  # Start from index 7
        ws.cell(row=r_idx, column=1, value=r_idx - 6)  # Add serial number starting from 1
        for c_idx, value in enumerate(row, start=2):  # Start from column 2 to avoid overwriting serial number
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Save the workbook to a BytesIO object
    from io import BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Serve the Excel file as an HTTP response
    response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=returned-products.xlsx'
    return response

    
    

class Order_return_Add(View):
    def get(self, request, route_id):
        template = 'order/order_return_add.html'
        form = Order_return_Form(route_id=route_id)
        
        return render(request, template, {'form': form})

    def post(self, request, route_id):
        template = 'order/order_return_add.html'
        form = Order_return_Form(route_id=route_id, data=request.POST)
        if form.is_valid():
            route = RouteMaster.objects.get(route_id=route_id)
            form.save(route=route)
            return redirect(order_return)
        customer = request.POST.get('customer')
        return render(request, template,)


class Order_return_Edit(View):
    template = 'order/order_return_edit.html'
    def get(self, request, order_return_id):
        customer = Order_return.objects.get(order_return_id=order_return_id).customer
        order_return = Order_return.objects.get(order_return_id=order_return_id)
        form = Order_return_Edit_Form(instance=order_return)
        return render(request, self.template, {'form': form, 'order_return': order_return, 'customer':customer})

    def post(self, request, order_return_id):
        order_return = Order_return.objects.get(order_return_id=order_return_id)
        form =Order_return_Edit_Form(request.POST, instance=order_return)
        
        if form.is_valid():
            form.save()
            return redirect('order_return')
        else:
            print(form.errors)
        return render(request, self.template, {'form': form, 'order_return': order_return})
    
class Order_Return_Delete(View):
    template='order/order_return_delete.html'
    def get(self, request, order_return_id):
        order_return = Order_return.objects.get(order_return_id=order_return_id)
        return render(request, self.template, {'order_return': order_return})

    def post(self, request, order_return_id):
        order_return = Order_return.objects.get(order_return_id=order_return_id)
        order_return.delete()
        return redirect('order_return')




# Order
def order_list(request):
    orders = Order.objects.all()
    
    return render(request, 'order/order_list.html', {'orders':orders})

class OrderCreate(View):
    def get(self, request):
        form = OrderForm()
        return render(request, 'order/order_create.html', {'form': form})

    def post(self, request):
        form = OrderForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('order_list')
        return render(request, 'order/order_create.html', {'form': form})


class OrderUpdate(View):
    def get(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        form = OrderUpdateForm(instance=order)
        return render(request, 'order/order_update.html', {'form': form})

    def post(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        form = OrderUpdateForm(request.POST, instance=order)
        if form.is_valid():
            order = form.save(commit=False)
            order.quantity = form.cleaned_data['quantity'] 
            order.order_date = form.cleaned_data['order_date']  # Update order date
            order.save()
            return redirect('order_list')
        return render(request, 'order/order_update.html', {'form': form})

class OrderDelete(View):
    def get(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        return render(request, 'order/order_delete.html', {'order': order})

    def post(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        order.delete()
        return redirect('order_list')
    
def order_by_date(request):
    selected_date = dt.today()
    if request.method == 'POST':
        selected_date = request.POST.get('date')
        if selected_date:  
            order_queryset = Order.objects.filter(order_date=selected_date)
        else:
            selected_date = dt.today()
            order_queryset = Order.objects.filter(order_date=selected_date)
    else:
        selected_date = dt.today()
        order_queryset = Order.objects.filter(order_date=selected_date)
    routes = RouteMaster.objects.all()
    route_saleman=[]
    for route in routes:
        route_sale={}
        order = order_queryset.filter(route=route).first()
        if order:
            salesman = order.salesman
            route_sale['route_id'] = route.route_id
            route_sale['route'] = route.route_name
            route_sale['salesman'] = salesman
            route_saleman.append(route_sale)
    context={'route_saleman':route_saleman, 'selected_date': selected_date}
    return render(request, 'order/order_by_date.html', context)


from datetime import datetime
def order_by_route(request, route_id, date, salesman_id):
    salesman =CustomUser.objects.get(id = salesman_id)
    date = datetime.strptime(date, '%Y-%m-%d').date()
    route=RouteMaster.objects.get(route_id=route_id)
    orders = Order.objects.filter(route=route, order_date=date)
    order_details = []
    for order in orders:
        order_detail={}
        stock = ProductStock.objects.filter(product_name = order.product.product_name).first()
        order_detail['stock'] = stock.quantity
        order_detail['product']=order.product.product_name
        order_detail['quantity']=order.quantity
        order_detail['unit']=order.product.unit
        found_duplicate = False
        for detail in order_details:
            if detail['product'] == order_detail['product']:
                found_duplicate = True
                # Update the quantity of the existing product entry
                detail['quantity'] += order_detail['quantity']
                print('Updated quantity for product:', order_detail['product'])
                break
        if not found_duplicate:
            order_details.append(order_detail)
    print(order_details)
    context={'route':route, 'date':date, 'order_details':order_details, 'salesman':salesman}
    return render (request, 'order/order_by_route.html', context)

def order_excel(request, route_id, date):
    return HttpResponse("Excel file generation in progress...")
